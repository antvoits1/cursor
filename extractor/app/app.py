from __future__ import annotations

import asyncio
import base64
import csv
import io
import hashlib
import html as html_lib
import ipaddress
import json
import os
import random
import re
import shutil
import socket
import sys
import webbrowser
import sqlite3
import tempfile
import uuid
import ssl
import time
from contextlib import asynccontextmanager, suppress
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any, AsyncGenerator, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from pathlib import Path
from urllib.parse import quote_plus, unquote, urljoin, urlparse
from urllib.request import Request as UrlRequest, urlopen

from curl_cffi.requests import AsyncSession
from email_validator import EmailNotValidError, validate_email
from fastapi import Body, FastAPI, Header, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
import phonenumbers
from phonenumbers import PhoneNumberMatcher, PhoneNumberType, carrier, geocoder, number_type
from selectolax.lexbor import LexborHTMLParser
import trafilatura

try:
    import dns.asyncresolver
    import dns.exception
    DNS_OK = True
except ImportError:
    DNS_OK = False

os.environ.setdefault("PLAYWRIGHT_BROWSERS_PATH", "0")

try:
    from playwright.async_api import Browser, Playwright, async_playwright
    PLAYWRIGHT_OK = True
except ImportError:
    Browser = Any  # type: ignore
    Playwright = Any  # type: ignore
    PLAYWRIGHT_OK = False

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_OK = True
except ImportError:
    Workbook = Any  # type: ignore
    load_workbook = Any  # type: ignore
    OPENPYXL_OK = False

try:
    from pypdf import PdfReader
    PYPDF_OK = True
except ImportError:
    PdfReader = Any  # type: ignore
    PYPDF_OK = False

try:
    from PIL import Image
    import pytesseract
    OCR_OK = True
except ImportError:
    Image = Any  # type: ignore
    pytesseract = Any  # type: ignore
    OCR_OK = False


# ---------------------------------------------------------------------------
# Application constants
# ---------------------------------------------------------------------------

FROZEN_BUILD = bool(getattr(sys, "frozen", False))
RESOURCE_DIR = os.path.dirname(os.path.abspath(sys.executable)) if FROZEN_BUILD else os.path.dirname(os.path.abspath(__file__))
if FROZEN_BUILD and os.name == "nt":
    DATA_DIR = os.path.join(os.environ.get("LOCALAPPDATA", RESOURCE_DIR), "IntelligenceExtractor", "Data")
else:
    DATA_DIR = RESOURCE_DIR
os.makedirs(DATA_DIR, exist_ok=True)
APP_DIR = RESOURCE_DIR
DB_PATH = os.path.join(DATA_DIR, "extractor_history.db")
ENGINE_VERSION = "2026.08.21-phase2-windows"
SCHEMA_VERSION = 2
DEFAULT_CACHE_TTL_DAYS = 7
DEFAULT_MAX_CONCURRENCY = 5
DEFAULT_HTTP_TIMEOUT = 16
DEFAULT_MAX_RESPONSE_BYTES = 6_000_000
MAX_REDIRECTS = 5
MAX_CUSTOM_URLS = 20
MAX_QUERY_LENGTH = 600
MAX_BULK_FILE_BYTES = 25_000_000
MAX_BULK_RECORDS = 500
MAX_SNAPSHOTS_PER_DOSSIER = 6
WEB_DIR = os.path.join(APP_DIR, "web")
PWA_DIR = os.path.join(APP_DIR, "pwa")
STATIC_DIR = os.path.join(APP_DIR, "static")
SNAPSHOT_DIR = os.path.join(DATA_DIR, "snapshots")
os.makedirs(SNAPSHOT_DIR, exist_ok=True)

if OCR_OK and os.name == "nt":
    bundled_tesseract = os.path.join(RESOURCE_DIR, "tesseract", "tesseract.exe")
    if os.path.isfile(bundled_tesseract):
        pytesseract.pytesseract.tesseract_cmd = bundled_tesseract
        tessdata_dir = os.path.join(RESOURCE_DIR, "tesseract", "tessdata")
        if os.path.isdir(tessdata_dir):
            os.environ.setdefault("TESSDATA_PREFIX", tessdata_dir)

GENERIC_EMAIL_PREFIXES = {
    "info", "sales", "support", "contact", "hello", "office", "admin", "service",
    "customerservice", "help", "billing", "accounts", "marketing", "team", "careers",
}
EMAIL_BLACKLIST_DOMAINS = {
    "example.com", "domain.com", "email.com", "sample.com", "test.com", "myshopify.com",
    "sentry.io", "wixpress.com", "wix.com", "wordpress.org", "wordpress.com", "schema.org",
    "cloudflare.com", "googleapis.com", "gravatar.com",
}
ASSET_SUBDOMAINS = {"media", "cdn", "assets", "static", "images", "img", "files", "downloads"}
OFFICER_ROLES = {
    "owner", "co-owner", "founder", "co-founder", "partner", "managing partner", "general partner",
    "president", "vice president", "ceo", "chief executive officer", "cfo", "chief financial officer",
    "coo", "chief operating officer", "cto", "chief technology officer", "principal", "director",
    "vice president", "vp", "regional sales director", "head",
    "managing member", "member", "manager", "officer", "chairman", "chairwoman", "registered agent",
}
RELATIONSHIP_LABELS = {"spouse", "wife", "husband", "relative", "relatives", "associate", "associates", "business associate", "business associates", "family member", "family", "partner"}
COMMON_MULTI_LABEL_PUBLIC_SUFFIXES = {
    "co.uk", "org.uk", "ac.uk", "gov.uk", "com.au", "net.au", "org.au", "co.nz", "com.br",
    "com.mx", "co.jp", "co.kr", "com.sg", "com.hk", "co.in", "com.cn", "com.tr", "com.ar",
}
US_STATES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR", "california": "CA",
    "colorado": "CO", "connecticut": "CT", "delaware": "DE", "florida": "FL", "georgia": "GA",
    "hawaii": "HI", "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV", "new hampshire": "NH",
    "new jersey": "NJ", "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR", "pennsylvania": "PA",
    "rhode island": "RI", "south carolina": "SC", "south dakota": "SD", "tennessee": "TN",
    "texas": "TX", "utah": "UT", "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
}
STATE_CODES = set(US_STATES.values())

SOURCE_PRIORITY = {
    "Official Website": 100,
    "BBB": 90,
    "D&B": 85,
    "Corporate Registry": 85,
    "PeopleSearchNow": 70,
    "USPhoneBook": 70,
    "TruePeopleSearch": 70,
    "FastPeopleSearch": 70,
    "That'sThem": 68,
    "NumLookup": 65,
    "Sync.ME": 65,
    "SpyDialer": 62,
    "Numpi": 60,
    "Search Discovery": 40,
    "Custom Target": 75,
}

SEARCH_ENGINES = (
    ("Bing RSS", "https://www.bing.com/search?format=rss&q={q}"),
    ("DuckDuckGo", "https://html.duckduckgo.com/html/?q={q}"),
    ("Bing", "https://www.bing.com/search?q={q}"),
    ("Yahoo", "https://search.yahoo.com/search?p={q}"),
    ("Google", "https://www.google.com/search?q={q}"),
)


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def iso_now() -> str:
    return utc_now().isoformat(timespec="seconds")


def clamp_int(value: Any, default: int, minimum: int, maximum: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, parsed))


def normalize_url(raw_url: str) -> str:
    text = (raw_url or "").strip()
    if not text:
        return ""
    if not re.match(r"^https?://", text, flags=re.I):
        text = "https://" + text
    return text


def clean_search_url(raw_url: str) -> str:
    raw_url = (raw_url or "").strip()
    if not raw_url:
        return ""
    if "uddg=" in raw_url:
        try:
            return unquote(raw_url.split("uddg=", 1)[1].split("&", 1)[0])
        except (ValueError, IndexError):
            return raw_url
    if "bing.com/ck/a" in raw_url:
        match = re.search(r"[?&]u=a1([A-Za-z0-9_\-+/=]+)", raw_url)
        if match:
            try:
                encoded = match.group(1)
                encoded += "=" * ((4 - len(encoded) % 4) % 4)
                decoded = base64.urlsafe_b64decode(encoded).decode("utf-8", errors="ignore")
                if decoded.startswith(("http://", "https://")):
                    return decoded
            except (ValueError, UnicodeDecodeError):
                return raw_url
    if "google.com/url" in raw_url:
        match = re.search(r"[?&]q=(https?://[^&]+)", raw_url)
        if match:
            return unquote(match.group(1))
    if "r.search.yahoo.com" in raw_url and "/RU=" in raw_url:
        try:
            return unquote(raw_url.split("/RU=", 1)[1].split("/RK=", 1)[0])
        except (ValueError, IndexError):
            return raw_url
    return raw_url


def hostname_from_url(raw_url: str) -> str:
    try:
        parsed = urlparse(normalize_url(raw_url))
        return (parsed.hostname or "").lower().rstrip(".")
    except ValueError:
        return ""


def registered_domain(hostname: str) -> str:
    host = (hostname or "").lower().strip(".")
    if not host:
        return ""
    labels = [x for x in host.split(".") if x]
    if len(labels) <= 2:
        return host
    last_two = ".".join(labels[-2:])
    if last_two in COMMON_MULTI_LABEL_PUBLIC_SUFFIXES and len(labels) >= 3:
        return ".".join(labels[-3:])
    return last_two


def domain_matches(candidate: str, target: str) -> bool:
    c = (candidate or "").lower().strip(".")
    t = (target or "").lower().strip(".")
    if not c or not t:
        return False
    return c == t or c.endswith("." + t)


def source_ref(label: str, url: str, **extra: Any) -> Dict[str, Any]:
    normalized = normalize_url(clean_search_url(url)) if url else ""
    host = hostname_from_url(normalized)
    out: Dict[str, Any] = {
        "label": label,
        "url": normalized,
        "domain": host,
        "captured_at": iso_now(),
    }
    out.update(extra)
    return out


def safe_company_case(name: str) -> str:
    legal = {"llc": "LLC", "inc": "INC", "ltd": "Ltd", "lp": "LP", "llp": "LLP", "pllc": "PLLC", "corp": "Corp", "co": "Co", "pc": "PC"}
    words = re.split(r"(\s+)", (name or "").strip())
    output: List[str] = []
    for token in words:
        if not token or token.isspace():
            output.append(token)
            continue
        match = re.match(r"^([^A-Za-z0-9]*)([A-Za-z0-9'\-]+)([^A-Za-z0-9]*)$", token)
        if not match:
            output.append(token)
            continue
        prefix, core, suffix = match.groups()
        lower = core.lower()
        if lower in legal:
            core_out = legal[lower]
        elif core.isupper() and 2 <= len(core) <= 8:
            core_out = core
        elif any(ch.isupper() for ch in core[1:]):
            core_out = core
        else:
            core_out = core[:1].upper() + core[1:]
        output.append(prefix + core_out + suffix)
    return "".join(output).strip()


def infer_company_from_domain(domain: str) -> str:
    root = registered_domain(domain)
    if not root:
        return ""
    stem = root.split(".", 1)[0]
    words = re.sub(r"[-_]+", " ", stem).strip()
    return safe_company_case(words)


def normalize_person_name(name: str) -> str:
    cleaned = re.sub(r"\b(?:mr|mrs|ms|miss|dr)\.?\s+", "", name or "", flags=re.I)
    cleaned = re.sub(r"[^A-Za-z0-9'\-\s]", " ", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


PERSON_TOKEN_PATTERN = r"[A-Z](?:[A-Za-z'\-]+|\.)?"
PERSON_NAME_PATTERN = rf"{PERSON_TOKEN_PATTERN}(?:\s+{PERSON_TOKEN_PATTERN}){{1,3}}"

PERSON_NAME_UI_WORDS = {
    "about", "address", "business", "call", "company", "contact", "details", "email", "information",
    "management", "message", "office", "phone", "profile", "rating", "review", "search", "support",
    "team", "website", "write", "customer", "customers", "principal", "contacts", "view", "visit",
}

def is_plausible_person_name(name: str) -> bool:
    cleaned = normalize_person_name(name)
    parts = cleaned.split()
    if not (2 <= len(parts) <= 4) or len(cleaned) > 80:
        return False
    lowered = {p.lower() for p in parts}
    if lowered & PERSON_NAME_UI_WORDS:
        return False
    if any(any(ch.isdigit() for ch in p) for p in parts):
        return False
    return all(re.fullmatch(r"[A-Za-z][A-Za-z'\-]*", p) is not None for p in parts)

def normalize_role(role: str) -> str:
    acronyms = {"ceo": "CEO", "cfo": "CFO", "coo": "COO", "cto": "CTO", "vp": "VP"}
    parts = re.split(r"(\s+)", (role or "").strip())
    return "".join(acronyms.get(part.lower(), part.capitalize()) if part and not part.isspace() else part for part in parts)

def person_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_person_name(name).lower())


def phone_key(raw: str) -> str:
    digits = re.sub(r"\D", "", raw or "")
    return digits[-10:] if len(digits) >= 10 else digits


def email_key(raw: str) -> str:
    return (raw or "").strip().lower()


def address_key(raw: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (raw or "").lower())


def csv_safe_value(value: Any) -> str:
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def decode_cfemail(hex_string: str) -> str:
    try:
        data = bytes.fromhex((hex_string or "").strip())
        if len(data) < 2:
            return ""
        key = data[0]
        return "".join(chr(b ^ key) for b in data[1:])
    except (ValueError, TypeError):
        return ""


def extract_cfemails(html: str) -> Set[str]:
    found: Set[str] = set()
    parser = LexborHTMLParser(html or "")
    for node in parser.root.css("[data-cfemail]"):
        decoded = decode_cfemail(node.attributes.get("data-cfemail", ""))
        if "@" in decoded:
            found.add(decoded)
    for node in parser.root.css("a[href*='/cdn-cgi/l/email-protection']"):
        href = node.attributes.get("href", "")
        match = re.search(r"#([0-9a-fA-F]+)$", href)
        if match:
            decoded = decode_cfemail(match.group(1))
            if "@" in decoded:
                found.add(decoded)
    return found


def extract_address_strings(text: str) -> List[str]:
    if not text:
        return []
    street_types = (
        r"Street|St\.?|Avenue|Ave\.?|Road|Rd\.?|Boulevard|Blvd\.?|Drive|Dr\.?|Lane|Ln\.?|"
        r"Highway|Hwy\.?|Way|Court|Ct\.?|Parkway|Pkwy\.?|Circle|Cir\.?|Trail|Trl\.?|"
        r"Terrace|Ter\.?|Place|Pl\.?|Plaza|Square|Sq\.?"
    )
    pattern = re.compile(
        rf"\b\d{{1,6}}\s+[A-Za-z0-9.'#\- ]{{2,70}}\s(?:{street_types})"
        rf"(?:\s+(?:Suite|Ste\.?|Unit|#)\s*[A-Za-z0-9\-]+)?"
        rf"(?:,?\s+[A-Za-z.'\- ]{{2,40}})?,?\s+[A-Z]{{2}}\s+\d{{5}}(?:-\d{{4}})?\b",
        re.I,
    )
    po_box = re.compile(
        r"\bP\.?\s*O\.?\s*Box\s+\d+[A-Za-z0-9\-]*,?\s+[A-Za-z.'\- ]{2,40},?\s+[A-Z]{2}\s+\d{5}(?:-\d{4})?\b",
        re.I,
    )
    results = [" ".join(m.group(0).split()) for m in pattern.finditer(text)]
    results.extend(" ".join(m.group(0).split()) for m in po_box.finditer(text))
    unique: List[str] = []
    seen: Set[str] = set()
    for item in results:
        key = address_key(item)
        if len(key) >= 10 and key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


def validate_external_http_url(raw_url: str) -> str:
    normalized = normalize_url(raw_url)
    parsed = urlparse(normalized)
    if parsed.scheme not in {"http", "https"}:
        raise ValueError("Only HTTP and HTTPS URLs are allowed")
    if not parsed.hostname:
        raise ValueError("URL must include a hostname")
    if parsed.username or parsed.password:
        raise ValueError("Credentials are not allowed in target URLs")
    return normalized


# ---------------------------------------------------------------------------
# Query parsing
# ---------------------------------------------------------------------------

class UniversalQueryParser:
    DOMAIN_PATTERN = re.compile(
        r"(?<!@)\b(?:https?://)?(?:www\.)?((?:[A-Za-z0-9](?:[A-Za-z0-9\-]{0,61}[A-Za-z0-9])?\.)+[A-Za-z]{2,63})(?:/[^\s]*)?",
        re.I,
    )
    EMAIL_PATTERN = re.compile(r"[A-Za-z0-9_.+\-]+@(?:[A-Za-z0-9\-]+\.)+[A-Za-z]{2,63}")

    @classmethod
    def parse(cls, raw_query: str) -> Dict[str, Any]:
        raw = (raw_query or "").strip()
        if not raw:
            raise ValueError("Query cannot be empty")
        if len(raw) > MAX_QUERY_LENGTH:
            raise ValueError(f"Query exceeds {MAX_QUERY_LENGTH} characters")

        work = raw
        emails = [m.group(0).lower() for m in cls.EMAIL_PATTERN.finditer(work)]
        inferred_domain = ""
        if emails:
            inferred_domain = emails[0].split("@", 1)[1].lower()
            work = cls.EMAIL_PATTERN.sub(" ", work)

        phones: List[str] = []
        masked = list(work)
        try:
            for match in PhoneNumberMatcher(work, "US"):
                formatted = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)
                if formatted not in phones:
                    phones.append(formatted)
                for idx in range(match.start, match.end):
                    if 0 <= idx < len(masked):
                        masked[idx] = " "
            work = "".join(masked)
        except (ValueError, TypeError, phonenumbers.NumberParseException):
            work = "".join(masked)

        domain = inferred_domain
        url_input = ""
        url_match = cls.DOMAIN_PATTERN.search(work)
        if url_match:
            token = url_match.group(0)
            url_input = token if token.lower().startswith(("http://", "https://")) else ""
            domain = registered_domain(url_match.group(1).lower())
            work = work[:url_match.start()] + " " + work[url_match.end():]

        owner, work = cls._extract_owner(work)
        state, work = cls._extract_state(work)

        company = re.sub(
            r"\b(?:company|business|corp(?:oration)?|information|details|lookup|research|search|find|contact)\b[:\-]?",
            " ",
            work,
            flags=re.I,
        )
        company = re.sub(r"\s+", " ", company).strip(" ,;:-")
        if not company and domain:
            company = infer_company_from_domain(domain)
        elif not company and emails:
            company = infer_company_from_domain(emails[0].split("@", 1)[1])
        company = safe_company_case(company)

        return {
            "raw": raw,
            "normalized_query": re.sub(r"\s+", " ", raw).strip(),
            "emails": emails,
            "phones": phones,
            "owner": normalize_person_name(owner),
            "state": state,
            "domain": domain,
            "url_input": url_input,
            "company_name": company,
        }

    @staticmethod
    def _extract_state(text: str) -> Tuple[str, str]:
        # State names/codes require location context, except an exact standalone state query.
        for full_name, code in sorted(US_STATES.items(), key=lambda x: len(x[0]), reverse=True):
            patterns = [
                re.compile(rf"\bin\s+{re.escape(full_name)}\b", re.I),
                re.compile(rf",\s*{re.escape(full_name)}\b", re.I),
                re.compile(rf"\bstate\s+of\s+{re.escape(full_name)}\b", re.I),
            ]
            if text.strip().lower() == full_name:
                return code, ""
            for pattern in patterns:
                match = pattern.search(text)
                if match:
                    return code, text[:match.start()] + " " + text[match.end():]

        contextual = re.search(r"(?:\bin\s+|,\s*)([A-Z]{2})\b", text)
        if contextual and contextual.group(1) in STATE_CODES:
            return contextual.group(1), text[:contextual.start()] + " " + text[contextual.end():]

        # A bare uppercase postal token is accepted only as a trailing location token.
        # This prevents company-name prefixes such as "ME Consulting" from becoming Maine.
        trailing = re.search(r"(?<![A-Za-z])([A-Z]{2})\s*$", text)
        if trailing and trailing.group(1) in STATE_CODES:
            return trailing.group(1), text[:trailing.start()] + " "
        return "", text

    @staticmethod
    def _extract_owner(text: str) -> Tuple[str, str]:
        role = r"owner|founder|co-founder|ceo|president|principal|partner|managing member|director|officer|attn|c/o"
        pattern = re.compile(
            rf"\b(?:{role})\b\s*[:\-]?\s+([A-Za-z][A-Za-z'\-]*(?:\s+[A-Za-z][A-Za-z'\-]*){{1,3}}?)(?=\s+(?:in|at|from)\s+(?:[A-Z]{{2}}|[A-Za-z ]+)|\s*,|\s*$)",
            re.I,
        )
        match = pattern.search(text)
        if not match:
            return "", text
        name = normalize_person_name(match.group(1))
        return name, text[:match.start()] + " " + text[match.end():]


# ---------------------------------------------------------------------------
# Relational dossier merge helpers
# ---------------------------------------------------------------------------

def make_empty_dossier(tokens: Dict[str, Any], cache_key: str) -> Dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "engine_version": ENGINE_VERSION,
        "cache_key": cache_key,
        "query": tokens["raw"],
        "entity": {
            "name": tokens["company_name"],
            "state": tokens["state"],
            "official_domain": tokens["domain"],
            "official_website": "",
            "phones": [],
            "emails": [],
            "addresses": [],
            "socials": [],
            "metadata": {},
        },
        "people": [],
        "relationships": [],
        "sources": [],
        "snapshots": [],
        "audit": {
            "started_at": iso_now(),
            "completed_at": "",
            "elapsed_seconds": 0.0,
            "snapshot_requests": {"attempted": 0, "captured": 0, "failed": 0},
            "requests": {
                "attempted": 0,
                "completed": 0,
                "successful": 0,
                "failed": 0,
                "blocked": 0,
                "challenged": 0,
                "rate_limited": 0,
            },
        },
    }


def evidence_key(src: Dict[str, Any]) -> Tuple[str, str]:
    return ((src.get("label") or "").lower(), (src.get("url") or "").lower())


def merge_evidence(target: List[Dict[str, Any]], additions: Iterable[Dict[str, Any]]) -> None:
    seen = {evidence_key(x) for x in target}
    for item in additions:
        key = evidence_key(item)
        if key not in seen:
            target.append(item)
            seen.add(key)


def merge_record_list(target: List[Dict[str, Any]], additions: Iterable[Dict[str, Any]], key_name: str) -> None:
    index: Dict[str, Dict[str, Any]] = {}
    for row in target:
        raw = str(row.get(key_name, ""))
        key = phone_key(raw) if key_name == "number" else email_key(raw) if key_name == "email" else address_key(raw)
        if key:
            index[key] = row
    for row in additions:
        raw = str(row.get(key_name, ""))
        key = phone_key(raw) if key_name == "number" else email_key(raw) if key_name == "email" else address_key(raw)
        if not key:
            continue
        if key not in index:
            target.append(row)
            index[key] = row
            continue
        existing = index[key]
        merge_evidence(existing.setdefault("sources", []), row.get("sources", []))
        for k, v in row.items():
            if k == "sources":
                continue
            if v not in (None, "", [], {}) and existing.get(k) in (None, "", [], {}, "unknown"):
                existing[k] = v


def merge_person(target_people: List[Dict[str, Any]], incoming: Dict[str, Any]) -> Dict[str, Any]:
    name = normalize_person_name(incoming.get("name", ""))
    key = person_key(name)
    if not key:
        return incoming
    existing = next((p for p in target_people if person_key(p.get("name", "")) == key), None)
    if existing is None:
        incoming["name"] = name
        incoming.setdefault("roles", [incoming.get("role", "")] if incoming.get("role") else [])
        incoming.setdefault("direct_phones", [])
        incoming.setdefault("direct_emails", [])
        incoming.setdefault("associated_addresses", [])
        incoming.setdefault("relatives_and_associates", [])
        incoming.setdefault("evidence_sources", [])
        target_people.append(incoming)
        return incoming

    new_role = (incoming.get("role") or "").strip()
    if new_role and new_role not in existing.setdefault("roles", []):
        existing["roles"].append(new_role)
    if not existing.get("role") and new_role:
        existing["role"] = new_role
    if not existing.get("corporate_relationship") and incoming.get("corporate_relationship"):
        existing["corporate_relationship"] = incoming["corporate_relationship"]
    merge_record_list(existing.setdefault("direct_phones", []), incoming.get("direct_phones", []), "number")
    merge_record_list(existing.setdefault("direct_emails", []), incoming.get("direct_emails", []), "email")
    merge_record_list(existing.setdefault("associated_addresses", []), incoming.get("associated_addresses", []), "address")
    merge_evidence(existing.setdefault("evidence_sources", []), incoming.get("evidence_sources", []))

    rel_index = {(r.get("name", "").lower(), r.get("relationship", "").lower()) for r in existing.setdefault("relatives_and_associates", [])}
    for rel in incoming.get("relatives_and_associates", []):
        rk = ((rel.get("name") or "").lower(), (rel.get("relationship") or "").lower())
        if rk not in rel_index:
            existing["relatives_and_associates"].append(rel)
            rel_index.add(rk)
    return existing


# ---------------------------------------------------------------------------
# Database layer
# ---------------------------------------------------------------------------

class Database:
    def __init__(self, path: str = DB_PATH):
        self.path = path
        self.fts_available = True

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.path, timeout=20)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA foreign_keys=ON;")
        return conn

    def initialize(self) -> None:
        conn = self.connect()
        try:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS app_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL DEFAULT ''
                )
            """)
            # Explicitly remove retired external-AI secrets/settings.
            conn.execute("DELETE FROM app_settings WHERE lower(key) LIKE '%api_key%'")
            conn.execute("""
                CREATE TABLE IF NOT EXISTS dossiers (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    cache_key TEXT NOT NULL UNIQUE,
                    query_text TEXT NOT NULL,
                    normalized_query TEXT NOT NULL,
                    company_name TEXT,
                    state TEXT,
                    phones_count INTEGER NOT NULL DEFAULT 0,
                    emails_count INTEGER NOT NULL DEFAULT 0,
                    dossier_json TEXT NOT NULL,
                    engine_version TEXT NOT NULL,
                    expires_at TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)
            conn.execute("CREATE INDEX IF NOT EXISTS idx_dossiers_updated ON dossiers(updated_at DESC)")
            conn.execute("CREATE INDEX IF NOT EXISTS idx_dossiers_company ON dossiers(company_name)")
            try:
                conn.execute("""
                    CREATE VIRTUAL TABLE IF NOT EXISTS dossiers_fts USING fts5(
                        cache_key UNINDEXED,
                        query_text,
                        company_name,
                        searchable_text
                    )
                """)
            except sqlite3.OperationalError:
                self.fts_available = False

            # Migrate legacy search_history once, then remove the obsolete table.
            legacy_exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='search_history'"
            ).fetchone()
            if legacy_exists:
                rows = conn.execute(
                    "SELECT query_text, company_name, state, phones_count, emails_count, dossier_json, created_at, updated_at FROM search_history"
                ).fetchall()
                for row in rows:
                    raw = row["query_text"] or ""
                    cache_key = hashlib.sha256(("legacy|" + raw).encode("utf-8")).hexdigest()
                    now = iso_now()
                    conn.execute(
                        """
                        INSERT OR IGNORE INTO dossiers(
                            cache_key, query_text, normalized_query, company_name, state,
                            phones_count, emails_count, dossier_json, engine_version,
                            expires_at, created_at, updated_at
                        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        (
                            cache_key, raw, re.sub(r"\s+", " ", raw).strip(), row["company_name"], row["state"],
                            row["phones_count"] or 0, row["emails_count"] or 0, row["dossier_json"] or "{}",
                            "legacy-import", now, str(row["created_at"] or now), str(row["updated_at"] or now),
                        ),
                    )
                conn.execute("DROP TABLE search_history")
            conn.commit()
        finally:
            conn.close()

    def get_setting(self, key: str, default: str = "") -> str:
        conn = self.connect()
        try:
            row = conn.execute("SELECT value FROM app_settings WHERE key=?", (key,)).fetchone()
            return row[0] if row else default
        finally:
            conn.close()

    def set_settings(self, values: Dict[str, str]) -> None:
        conn = self.connect()
        try:
            for key, value in values.items():
                conn.execute(
                    "INSERT INTO app_settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (key, value),
                )
            conn.commit()
        finally:
            conn.close()

    def cache_key(self, normalized_query: str, custom_urls: Sequence[str], proxy: str) -> str:
        source_signature = {
            "sources": [
                "official", "bbb", "dnb", "peoplesearchnow", "usphonebook", "truepeoplesearch",
                "fastpeoplesearch", "thatsthem", "numlookup", "syncme", "spydialer", "numpi",
            ],
            "custom_urls": sorted(custom_urls),
            "proxy": proxy,
            "engine": ENGINE_VERSION,
        }
        payload = json.dumps([normalized_query.lower(), source_signature], sort_keys=True, separators=(",", ":"))
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def get_cached(self, cache_key: str) -> Optional[Dict[str, Any]]:
        conn = self.connect()
        try:
            row = conn.execute("SELECT dossier_json, expires_at FROM dossiers WHERE cache_key=?", (cache_key,)).fetchone()
            if not row:
                return None
            expires = row["expires_at"]
            if expires:
                try:
                    if datetime.fromisoformat(expires) <= utc_now():
                        return None
                except ValueError:
                    return None
            try:
                dossier = json.loads(row["dossier_json"])
            except json.JSONDecodeError:
                return None
            if dossier.get("schema_version") != SCHEMA_VERSION:
                return None
            return dossier
        finally:
            conn.close()

    def get_by_cache_key(self, cache_key: str) -> Optional[Dict[str, Any]]:
        if not cache_key:
            return None
        conn = self.connect()
        try:
            row = conn.execute("SELECT dossier_json FROM dossiers WHERE cache_key=?", (cache_key,)).fetchone()
            if not row:
                return None
            try:
                dossier = json.loads(row[0])
            except json.JSONDecodeError:
                return None
            return dossier if dossier.get("schema_version") == SCHEMA_VERSION else None
        finally:
            conn.close()

    def save_dossier(self, dossier: Dict[str, Any], ttl_days: int) -> None:
        cache_key = dossier["cache_key"]
        entity = dossier.get("entity", {})
        query_text = dossier.get("query", "")
        normalized = re.sub(r"\s+", " ", query_text).strip()
        expires = (utc_now() + timedelta(days=ttl_days)).isoformat(timespec="seconds")
        now = iso_now()
        phones_count = len(entity.get("phones", [])) + sum(len(p.get("direct_phones", [])) for p in dossier.get("people", []))
        emails_count = len(entity.get("emails", [])) + sum(len(p.get("direct_emails", [])) for p in dossier.get("people", []))
        blob = json.dumps(dossier, ensure_ascii=False)
        searchable_parts = [query_text, entity.get("name", ""), entity.get("official_domain", "")]
        for p in dossier.get("people", []):
            searchable_parts.extend([p.get("name", ""), p.get("role", "")])
            searchable_parts.extend(x.get("number", "") for x in p.get("direct_phones", []))
            searchable_parts.extend(x.get("email", "") for x in p.get("direct_emails", []))
            searchable_parts.extend(x.get("address", "") for x in p.get("associated_addresses", []))
            searchable_parts.extend(x.get("name", "") for x in p.get("relatives_and_associates", []))
        searchable_parts.extend(x.get("number", "") for x in entity.get("phones", []))
        searchable_parts.extend(x.get("email", "") for x in entity.get("emails", []))
        searchable_parts.extend(x.get("address", "") for x in entity.get("addresses", []))
        searchable_text = " ".join(x for x in searchable_parts if x)

        conn = self.connect()
        try:
            conn.execute(
                """
                INSERT INTO dossiers(
                    cache_key, query_text, normalized_query, company_name, state, phones_count,
                    emails_count, dossier_json, engine_version, expires_at, created_at, updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(cache_key) DO UPDATE SET
                    query_text=excluded.query_text,
                    normalized_query=excluded.normalized_query,
                    company_name=excluded.company_name,
                    state=excluded.state,
                    phones_count=excluded.phones_count,
                    emails_count=excluded.emails_count,
                    dossier_json=excluded.dossier_json,
                    engine_version=excluded.engine_version,
                    expires_at=excluded.expires_at,
                    updated_at=excluded.updated_at
                """,
                (
                    cache_key, query_text, normalized, entity.get("name", ""), entity.get("state", ""),
                    phones_count, emails_count, blob, ENGINE_VERSION, expires, now, now,
                ),
            )
            if self.fts_available:
                conn.execute("DELETE FROM dossiers_fts WHERE cache_key=?", (cache_key,))
                conn.execute(
                    "INSERT INTO dossiers_fts(cache_key, query_text, company_name, searchable_text) VALUES(?,?,?,?)",
                    (cache_key, query_text, entity.get("name", ""), searchable_text),
                )
            conn.commit()
        finally:
            conn.close()

    def history(self, limit: int = 60) -> List[Dict[str, Any]]:
        conn = self.connect()
        try:
            rows = conn.execute(
                """
                SELECT id, cache_key, query_text, company_name, state, phones_count, emails_count, engine_version, updated_at
                FROM dossiers ORDER BY updated_at DESC LIMIT ?
                """,
                (limit,),
            ).fetchall()
            return [dict(r) for r in rows]
        finally:
            conn.close()

    def history_item(self, history_id: int) -> Optional[Dict[str, Any]]:
        conn = self.connect()
        try:
            row = conn.execute("SELECT dossier_json FROM dossiers WHERE id=?", (history_id,)).fetchone()
            if not row:
                return None
            try:
                return json.loads(row[0])
            except json.JSONDecodeError:
                return None
        finally:
            conn.close()

    def latest_for_entity(self, entity_name: str) -> Optional[Dict[str, Any]]:
        if not entity_name:
            return None
        conn = self.connect()
        try:
            row = conn.execute(
                "SELECT dossier_json FROM dossiers WHERE lower(company_name)=lower(?) ORDER BY updated_at DESC LIMIT 1",
                (entity_name,),
            ).fetchone()
            if not row:
                return None
            try:
                return json.loads(row[0])
            except json.JSONDecodeError:
                return None
        finally:
            conn.close()

    def delete_history(self, history_id: int) -> bool:
        conn = self.connect()
        try:
            row = conn.execute("SELECT cache_key FROM dossiers WHERE id=?", (history_id,)).fetchone()
            if not row:
                return False
            cache_key = row[0]
            conn.execute("DELETE FROM dossiers WHERE id=?", (history_id,))
            if self.fts_available:
                conn.execute("DELETE FROM dossiers_fts WHERE cache_key=?", (cache_key,))
            conn.commit()
            return True
        finally:
            conn.close()

    def clear_history(self) -> None:
        conn = self.connect()
        try:
            conn.execute("DELETE FROM dossiers")
            if self.fts_available:
                conn.execute("DELETE FROM dossiers_fts")
            conn.commit()
        finally:
            conn.close()


DB = Database()


# ---------------------------------------------------------------------------
# Network safety, pooling, and telemetry
# ---------------------------------------------------------------------------

@dataclass
class RequestMetrics:
    attempted: int = 0
    completed: int = 0
    successful: int = 0
    failed: int = 0
    blocked: int = 0
    challenged: int = 0
    rate_limited: int = 0

    def as_dict(self) -> Dict[str, int]:
        return {
            "attempted": self.attempted,
            "completed": self.completed,
            "successful": self.successful,
            "failed": self.failed,
            "blocked": self.blocked,
            "challenged": self.challenged,
            "rate_limited": self.rate_limited,
        }


@dataclass
class FetchResult:
    url: str
    final_url: str
    html: str = ""
    status: int = 0
    tier: str = ""
    duration: float = 0.0
    event: str = "SOURCE_FAILED"
    error: str = ""


class SSRFGuard:
    @staticmethod
    async def validate(raw_url: str) -> str:
        normalized = validate_external_http_url(raw_url)
        host = hostname_from_url(normalized)
        if host in {"localhost", "localhost.localdomain"}:
            raise ValueError("Localhost targets are blocked")
        literal = None
        try:
            literal = ipaddress.ip_address(host)
        except ValueError:
            literal = None
        if literal is not None:
            SSRFGuard._reject_ip(literal)
            return normalized

        try:
            records = await asyncio.to_thread(socket.getaddrinfo, host, None, type=socket.SOCK_STREAM)
        except socket.gaierror as exc:
            raise ValueError(f"Unable to resolve target host: {host}") from exc
        if not records:
            raise ValueError("Target host did not resolve")
        for record in records:
            ip_text = record[4][0]
            try:
                SSRFGuard._reject_ip(ipaddress.ip_address(ip_text))
            except ValueError as exc:
                raise ValueError(f"Unsafe target address: {ip_text}") from exc
        return normalized

    @staticmethod
    def _reject_ip(ip: ipaddress._BaseAddress) -> None:
        if (
            ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_multicast or ip.is_reserved
            or ip.is_unspecified
        ):
            raise ValueError(f"Private/local/reserved target blocked: {ip}")
        metadata = ipaddress.ip_address("169.254.169.254")
        if ip == metadata:
            raise ValueError("Cloud metadata address is blocked")


def playwright_proxy_config(proxy_url: str) -> Optional[Dict[str, str]]:
    if not proxy_url:
        return None
    parsed = urlparse(proxy_url)
    if not parsed.scheme or not parsed.hostname:
        return None
    server = f"{parsed.scheme}://{parsed.hostname}" + (f":{parsed.port}" if parsed.port else "")
    config: Dict[str, str] = {"server": server}
    if parsed.username:
        config["username"] = unquote(parsed.username)
    if parsed.password:
        config["password"] = unquote(parsed.password)
    return config


class BrowserPool:
    def __init__(self) -> None:
        self._lock = asyncio.Lock()
        self._playwright: Optional[Playwright] = None
        self._browser: Optional[Browser] = None
        self._unavailable_reason = ""

    async def ensure_browser(self) -> Optional[Browser]:
        if not PLAYWRIGHT_OK or self._unavailable_reason:
            return None
        if self._browser is not None:
            return self._browser
        async with self._lock:
            if self._browser is None:
                try:
                    self._playwright = await async_playwright().start()
                    self._browser = await self._playwright.chromium.launch(headless=True)
                except Exception as default_error:
                    # Fall back to a system Chromium/Chrome binary before declaring snapshots unavailable.
                    system_browser = next((path for path in (shutil.which("chromium"), shutil.which("chromium-browser"), shutil.which("google-chrome"), shutil.which("google-chrome-stable")) if path), None)
                    if self._playwright is not None and system_browser:
                        try:
                            self._browser = await self._playwright.chromium.launch(headless=True, executable_path=system_browser, args=["--no-sandbox"])
                            return self._browser
                        except Exception:
                            self._browser = None
                    if self._playwright is not None:
                        with suppress(Exception):
                            await self._playwright.stop()
                    self._playwright = None
                    self._browser = None
                    self._unavailable_reason = f"Chromium browser executable unavailable: {str(default_error)[:140]}"
                    return None
        return self._browser

    async def capture_snapshot(self, url: str, proxy: str, label: str, cache_key: str) -> Tuple[Optional[Dict[str, Any]], str]:
        browser = await self.ensure_browser()
        if browser is None:
            return None, self._unavailable_reason or "Playwright unavailable"
        context = None
        started = time.perf_counter()
        try:
            safe_url = await SSRFGuard.validate(url)
            context_args: Dict[str, Any] = {
                "ignore_https_errors": False,
                "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
                "viewport": {"width": 1280, "height": 820},
            }
            proxy_config = playwright_proxy_config(proxy)
            if proxy_config:
                context_args["proxy"] = proxy_config
            context = await browser.new_context(**context_args)
            await context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined});")
            page = await context.new_page()
            response = await page.goto(safe_url, wait_until="domcontentloaded", timeout=DEFAULT_HTTP_TIMEOUT * 1000)
            status = response.status if response else 0
            if status >= 400:
                return None, f"Snapshot page returned HTTP {status}"
            safe_label = re.sub(r"[^a-z0-9]+", "-", label.lower()).strip("-")[:32] or "source"
            fingerprint = hashlib.sha1(safe_url.encode("utf-8")).hexdigest()[:10]
            filename = f"{cache_key[:12] or 'direct'}-{safe_label}-{fingerprint}.jpg"
            full_path = os.path.join(SNAPSHOT_DIR, filename)
            await page.screenshot(path=full_path, type="jpeg", quality=58, full_page=False)
            return {
                "label": label,
                "source_url": page.url,
                "image_url": f"/snapshots/{filename}",
                "captured_at": iso_now(),
                "duration_seconds": round(time.perf_counter() - started, 2),
            }, ""
        except Exception as exc:
            return None, str(exc)[:240]
        finally:
            if context is not None:
                with suppress(Exception):
                    await context.close()

    async def fetch(self, url: str, proxy: str, timeout_ms: int, max_bytes: int) -> FetchResult:
        started = time.perf_counter()
        browser = await self.ensure_browser()
        if browser is None:
            return FetchResult(url=url, final_url=url, tier="Playwright", duration=0.0, error="Playwright unavailable")
        context = None
        try:
            context_args: Dict[str, Any] = {
                "ignore_https_errors": False,
                "user_agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36",
                "viewport": {"width": 1440, "height": 950},
            }
            proxy_config = playwright_proxy_config(proxy)
            if proxy_config:
                context_args["proxy"] = proxy_config
            context = await browser.new_context(**context_args)
            await context.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined});")
            page = await context.new_page()
            response = await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            status = response.status if response else 0
            final_url = page.url
            content = await page.content()
            if len(content.encode("utf-8", errors="ignore")) > max_bytes:
                content = content[:max_bytes]
            event = detect_network_event(status, content)
            return FetchResult(
                url=url, final_url=final_url, html=content, status=status, tier="Playwright",
                duration=round(time.perf_counter() - started, 3), event=event,
            )
        except Exception as exc:
            return FetchResult(
                url=url, final_url=url, tier="Playwright", duration=round(time.perf_counter() - started, 3),
                event="SOURCE_FAILED", error=str(exc)[:300],
            )
        finally:
            if context is not None:
                await context.close()

    async def close(self) -> None:
        async with self._lock:
            if self._browser is not None:
                await self._browser.close()
                self._browser = None
            if self._playwright is not None:
                await self._playwright.stop()
                self._playwright = None


BROWSER_POOL = BrowserPool()


def detect_network_event(status: int, html: str) -> str:
    lower = (html or "").lower()
    if status == 429:
        return "RATE_LIMITED_429"
    if status == 403:
        return "CHALLENGE_DETECTED" if any(x in lower for x in ("cloudflare", "turnstile", "verify you are human", "challenge-platform")) else "SOURCE_BLOCKED_403"
    if any(x in lower for x in ("cf-turnstile", "challenge-platform", "attention required! | cloudflare", "verify you are human")):
        return "CHALLENGE_DETECTED"
    if 200 <= status < 400:
        return "SOURCE_COMPLETED"
    return "SOURCE_FAILED"


class NetworkClient:
    def __init__(self, proxy: str = "", max_concurrency: int = DEFAULT_MAX_CONCURRENCY):
        self.proxy = proxy.strip()
        self.metrics = RequestMetrics()
        self.session: Optional[AsyncSession] = None
        self.semaphore = asyncio.Semaphore(max_concurrency)
        self.timeout = DEFAULT_HTTP_TIMEOUT
        self.max_bytes = DEFAULT_MAX_RESPONSE_BYTES

    async def __aenter__(self) -> "NetworkClient":
        self.session = AsyncSession(impersonate="chrome124")
        return self

    async def __aexit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        if self.session is not None:
            await self.session.close()
            self.session = None

    async def fetch(self, raw_url: str, *, allow_browser: bool = True, retries: int = 2, validate_ssrf: bool = True) -> FetchResult:
        async with self.semaphore:
            started = time.perf_counter()
            try:
                current = await SSRFGuard.validate(raw_url) if validate_ssrf else validate_external_http_url(raw_url)
            except ValueError as exc:
                return FetchResult(url=raw_url, final_url=raw_url, event="SOURCE_FAILED", error=str(exc))

            if self.session is None:
                self.session = AsyncSession(impersonate="chrome124")

            http_result = FetchResult(url=current, final_url=current)
            redirect_count = 0
            attempt = 0
            while attempt <= retries:
                self.metrics.attempted += 1
                try:
                    response = await self.session.get(
                        current, timeout=self.timeout, allow_redirects=False, proxy=self.proxy or None,
                        headers={"Accept-Language": "en-US,en;q=0.9"},
                    )
                    status = int(response.status_code)
                    if status in {301, 302, 303, 307, 308}:
                        redirect_result = FetchResult(url=raw_url, final_url=current, status=status, tier="curl_cffi", event="SOURCE_COMPLETED")
                        self._count_result(redirect_result)
                        location = response.headers.get("location", "")
                        if not location or redirect_count >= MAX_REDIRECTS:
                            return FetchResult(url=raw_url, final_url=current, status=status, tier="curl_cffi", event="SOURCE_FAILED", error="Invalid or excessive redirect chain")
                        target = urljoin(current, location)
                        current = await SSRFGuard.validate(target) if validate_ssrf else validate_external_http_url(target)
                        redirect_count += 1
                        continue

                    content = response.text or ""
                    if len(content.encode("utf-8", errors="ignore")) > self.max_bytes:
                        content = content[: self.max_bytes]
                    event = detect_network_event(status, content)
                    http_result = FetchResult(
                        url=raw_url, final_url=current, html=content, status=status, tier="curl_cffi",
                        duration=round(time.perf_counter() - started, 3), event=event,
                    )
                    self._count_result(http_result)
                    if (event in {"RATE_LIMITED_429", "SOURCE_BLOCKED_403"} or status == 503) and attempt < retries:
                        await asyncio.sleep((2 ** attempt) + random.uniform(0.25, 1.25))
                        attempt += 1
                        continue
                    break
                except Exception as exc:
                    http_result = FetchResult(
                        url=raw_url, final_url=current, tier="curl_cffi",
                        duration=round(time.perf_counter() - started, 3), event="SOURCE_FAILED", error=str(exc)[:300],
                    )
                    self._count_result(http_result)
                    if attempt < retries:
                        await asyncio.sleep((2 ** attempt) + random.uniform(0.25, 1.0))
                        attempt += 1
                        continue
                    break

            if allow_browser and http_result.event in {"CHALLENGE_DETECTED", "SOURCE_BLOCKED_403", "SOURCE_FAILED"} and PLAYWRIGHT_OK:
                self.metrics.attempted += 1
                browser_result = await BROWSER_POOL.fetch(current, self.proxy, self.timeout * 1000, self.max_bytes)
                self._count_result(browser_result)
                if browser_result.event == "SOURCE_COMPLETED":
                    return browser_result
            return http_result

    def _count_result(self, result: FetchResult) -> None:
        if result.status > 0:
            self.metrics.completed += 1
        if result.event == "SOURCE_COMPLETED":
            self.metrics.successful += 1
        elif result.event == "RATE_LIMITED_429":
            self.metrics.rate_limited += 1
            self.metrics.failed += 1
        elif result.event == "SOURCE_BLOCKED_403":
            self.metrics.blocked += 1
            self.metrics.failed += 1
        elif result.event == "CHALLENGE_DETECTED":
            self.metrics.challenged += 1
            self.metrics.failed += 1
        else:
            self.metrics.failed += 1


# ---------------------------------------------------------------------------
# Contact extraction and verification
# ---------------------------------------------------------------------------

class ContactVerifier:
    def __init__(self, smtp_checks: bool = True):
        self.smtp_checks = smtp_checks
        self._mx_cache: Dict[str, Dict[str, Any]] = {}

    async def verify_email(self, email: str, target_domain: str) -> Dict[str, Any]:
        normalized = email_key(email)
        try:
            valid = validate_email(normalized, check_deliverability=False)
            normalized = valid.normalized
            syntax_valid = True
        except EmailNotValidError:
            return {
                "email": normalized,
                "syntax_valid": False,
                "domain_match": False,
                "mx_status": "invalid_syntax",
                "smtp_status": "not_checked",
            }
        domain = normalized.split("@", 1)[1]
        mx = await self.lookup_mx(domain)
        smtp_status = "not_checked"
        if self.smtp_checks and mx.get("status") == "active":
            smtp_status = await self.smtp_probe(normalized, mx.get("hosts", []))
        return {
            "email": normalized,
            "syntax_valid": syntax_valid,
            "domain_match": domain_matches(domain, target_domain),
            "mx_status": mx.get("status", "unknown"),
            "smtp_status": smtp_status,
        }

    async def lookup_mx(self, domain: str) -> Dict[str, Any]:
        domain = (domain or "").lower().strip(".")
        if domain in self._mx_cache:
            return self._mx_cache[domain]
        if not DNS_OK:
            result = {"status": "unknown", "hosts": [], "reason": "dnspython unavailable"}
            self._mx_cache[domain] = result
            return result
        try:
            resolver = dns.asyncresolver.Resolver()
            resolver.lifetime = 5
            answers = await resolver.resolve(domain, "MX")
            hosts = [str(r.exchange).rstrip(".") for r in sorted(answers, key=lambda r: int(r.preference))]
            result = {"status": "active" if hosts else "no_mail_server", "hosts": hosts}
        except Exception as exc:
            name = exc.__class__.__name__.lower()
            status = "no_mail_server" if "noanswer" in name or "nxdomain" in name else "unknown"
            result = {"status": status, "hosts": [], "reason": str(exc)[:180]}
        self._mx_cache[domain] = result
        return result

    async def smtp_probe(self, email: str, mx_hosts: Sequence[str]) -> str:
        if not mx_hosts:
            return "no_mail_server"
        return await asyncio.to_thread(self._smtp_probe_sync, email, list(mx_hosts)[:2])

    @staticmethod
    def _smtp_probe_sync(email: str, mx_hosts: Sequence[str]) -> str:
        import smtplib
        for host in mx_hosts:
            try:
                with smtplib.SMTP(host, 25, timeout=6) as smtp:
                    smtp.ehlo_or_helo_if_needed()
                    smtp.mail("")
                    code, _ = smtp.rcpt(email)
                    if code in {250, 251, 252}:
                        domain = email.split("@", 1)[1]
                        probe_local = "extractor-probe-" + hashlib.sha1(str(time.time_ns()).encode("ascii")).hexdigest()[:10]
                        probe_code, _ = smtp.rcpt(f"{probe_local}@{domain}")
                        return "catch_all" if probe_code in {250, 251, 252} else "active_inbox"
                    if 500 <= code < 600:
                        return "rejected"
            except (OSError, smtplib.SMTPException, ssl.SSLError):
                continue
        return "unknown"


def classify_phone(raw: str, label: str, url: str) -> Optional[Dict[str, Any]]:
    try:
        parsed = phonenumbers.parse(raw, "US")
    except phonenumbers.NumberParseException:
        return None
    if not phonenumbers.is_possible_number(parsed) or not phonenumbers.is_valid_number(parsed):
        return None
    ptype = number_type(parsed)
    if ptype == PhoneNumberType.MOBILE:
        line_type = "Mobile"
    elif ptype == PhoneNumberType.FIXED_LINE:
        line_type = "Landline"
    elif ptype == PhoneNumberType.VOIP:
        line_type = "VoIP"
    elif ptype == PhoneNumberType.TOLL_FREE:
        line_type = "Toll-Free"
    else:
        line_type = "Fixed/Mobile"
    e164 = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
    national = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.NATIONAL)
    return {
        "number": national,
        "e164": e164,
        "digits": phone_key(e164),
        "line_type": line_type,
        "carrier": carrier.name_for_number(parsed, "en") or "Unknown",
        "region": geocoder.description_for_number(parsed, "en") or "",
        "sources": [source_ref(label, url)],
    }


async def extract_emails(html: str, text: str, target_domain: str, label: str, url: str, verifier: ContactVerifier) -> List[Dict[str, Any]]:
    candidates: Set[str] = set(extract_cfemails(html))
    parser = LexborHTMLParser(html or "")
    for a in parser.root.css("a[href]"):
        href = a.attributes.get("href", "")
        if href.lower().startswith("mailto:"):
            value = href.split(":", 1)[1].split("?", 1)[0]
            for item in value.split(","):
                if "@" in item:
                    candidates.add(unquote(item).strip())
    pattern = re.compile(r"[A-Za-z0-9_.+\-]+@(?:[A-Za-z0-9\-]+\.)+[A-Za-z]{2,63}")
    candidates.update(pattern.findall((html or "") + " " + (text or "")))

    results: List[Dict[str, Any]] = []
    for candidate in sorted(candidates):
        lowered = candidate.lower().strip().rstrip(".,;:")
        domain = lowered.split("@")[-1] if "@" in lowered else ""
        if domain in EMAIL_BLACKLIST_DOMAINS or lowered.endswith((".png", ".jpg", ".gif", ".svg", ".css", ".js")):
            continue
        status = await verifier.verify_email(lowered, target_domain)
        if not status.get("syntax_valid"):
            continue
        status["sources"] = [source_ref(label, url)]
        results.append(status)
    return results


def extract_phones(html: str, text: str, label: str, url: str) -> List[Dict[str, Any]]:
    candidates: Set[str] = set()
    parser = LexborHTMLParser(html or "")
    for a in parser.root.css("a[href]"):
        href = a.attributes.get("href", "")
        if href.lower().startswith("tel:"):
            candidates.add(unquote(href.split(":", 1)[1].split("?", 1)[0]))
    for match in PhoneNumberMatcher(text or "", "US"):
        candidates.add(match.raw_string)
    results: List[Dict[str, Any]] = []
    seen_digits: Set[str] = set()
    for raw in candidates:
        item = classify_phone(raw, label, url)
        if item and item.get("digits") not in seen_digits:
            seen_digits.add(item.get("digits", ""))
            results.append(item)
    return results


def extract_addresses(html: str, text: str, label: str, url: str) -> List[Dict[str, Any]]:
    parser = LexborHTMLParser(html or "")
    candidates: List[str] = []
    for node in parser.root.css("address, [class*='address'], [data-address], .location, footer"):
        candidates.extend(extract_address_strings(node.text(separator=" ")))
    candidates.extend(extract_address_strings(text))
    unique: List[Dict[str, Any]] = []
    seen: Set[str] = set()
    for address in candidates:
        key = address_key(address)
        if key not in seen:
            seen.add(key)
            unique.append({
                "address": address,
                "sources": [source_ref(label, url)],
                "maps_url": f"https://www.google.com/maps/search/?api=1&query={quote_plus(address)}",
            })
    return unique


def extract_socials(html: str, label: str, url: str) -> List[Dict[str, Any]]:
    parser = LexborHTMLParser(html or "")
    results: List[Dict[str, Any]] = []
    seen: Set[str] = set()
    for a in parser.root.css("a[href]"):
        href = clean_search_url(a.attributes.get("href", ""))
        low = href.lower()
        if href.startswith("http") and any(x in low for x in ("linkedin.com/", "facebook.com/", "instagram.com/", "x.com/", "twitter.com/", "youtube.com/")):
            if href not in seen:
                seen.add(href)
                results.append({"url": href, "sources": [source_ref(label, url)]})
    return results


def parse_jsonld_people_addresses(html: str, label: str, url: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    parser = LexborHTMLParser(html or "")
    people: List[Dict[str, Any]] = []
    addresses: List[Dict[str, Any]] = []

    def walk(node: Any) -> Iterable[Dict[str, Any]]:
        if isinstance(node, dict):
            yield node
            for value in node.values():
                yield from walk(value)
        elif isinstance(node, list):
            for value in node:
                yield from walk(value)

    for script in parser.root.css('script[type="application/ld+json"]'):
        try:
            raw = json.loads(script.text())
        except json.JSONDecodeError:
            continue
        for node in walk(raw):
            addr = node.get("address")
            addr_items = addr if isinstance(addr, list) else [addr] if addr else []
            for item in addr_items:
                if isinstance(item, dict):
                    parts = [item.get("streetAddress", ""), item.get("addressLocality", ""), item.get("addressRegion", ""), item.get("postalCode", "")]
                    address = ", ".join(str(x).strip() for x in parts if x)
                    if address:
                        addresses.append({"address": address, "sources": [source_ref(f"{label} Schema", url)]})
            for key in ("founder", "founders", "owner", "member"):
                raw_people = node.get(key)
                if not raw_people:
                    continue
                items = raw_people if isinstance(raw_people, list) else [raw_people]
                for person in items:
                    if isinstance(person, dict):
                        name = person.get("name", "")
                    elif isinstance(person, str):
                        name = person
                    else:
                        name = ""
                    name = normalize_person_name(str(name))
                    if len(name.split()) >= 2:
                        people.append({
                            "name": name,
                            "role": normalize_role(key.replace("s", "")),
                            "corporate_relationship": normalize_role(key.replace("s", "")),
                            "direct_phones": [], "direct_emails": [], "associated_addresses": [],
                            "relatives_and_associates": [],
                            "evidence_sources": [source_ref(f"{label} Schema", url)],
                        })
    return people, addresses


def extract_people_from_text(text: str, label: str, url: str, allow_unseparated: bool = False) -> List[Dict[str, Any]]:
    source_text = text or ""
    results: List[Dict[str, Any]] = []
    role_pattern = "|".join(sorted((re.escape(r) for r in OFFICER_ROLES), key=len, reverse=True))
    role_group = rf"(?i:{role_pattern})"
    title = r"(?:(?i:Mr|Ms|Mrs|Dr)\.?\s+)?"
    name = title + rf"({PERSON_TOKEN_PATTERN}(?:\s+{PERSON_TOKEN_PATTERN}){{1,3}}?)"
    patterns = [
        (re.compile(rf"{name}\s*[-–—,:|]\s*({role_group})\b"), "name_role"),
        (re.compile(rf"\b({role_group})\b\s*[-–—,:|]\s*{name}"), "role_name"),
    ]
    # Separator-free forms are useful for compact team cards ("Leo Chang CEO") but
    # are intentionally disabled on flattened full-page text so one card cannot bleed into the next.
    if allow_unseparated:
        patterns.append((re.compile(rf"{name}\s+({role_group})\b"), "name_role"))
    seen: Set[Tuple[str, str]] = set()
    for pattern, orientation in patterns:
        for match in pattern.finditer(source_text):
            if orientation == "name_role":
                person_name, role = match.group(1), match.group(2)
            else:
                role, person_name = match.group(1), match.group(2)
            person_name = normalize_person_name(person_name)
            if not is_plausible_person_name(person_name):
                continue
            role = normalize_role(role)
            key = (person_key(person_name), role.lower())
            if key in seen:
                continue
            seen.add(key)
            results.append({
                "name": person_name,
                "role": role,
                "corporate_relationship": role,
                "direct_phones": [], "direct_emails": [], "associated_addresses": [],
                "relatives_and_associates": [],
                "evidence_sources": [source_ref(label, url)],
            })
    return results


def extract_labeled_people(html: str, label: str, url: str) -> List[Dict[str, Any]]:
    parser = LexborHTMLParser(html or "")
    people: List[Dict[str, Any]] = []
    role_rx = re.compile(r"\b(owner|principal|president|ceo|chief executive officer|founder|co-founder|managing member|manager|partner|director|officer|business management|registered agent|vice president|vp|regional sales director|head)\b", re.I)
    for node in parser.root.css("dt, dd, li, p, div, section, article, tr"):
        text = re.sub(r"\s+", " ", node.text(separator=" ")).strip()
        if not text or len(text) > 700 or not role_rx.search(text):
            continue
        role_hits = role_rx.findall(text)
        people.extend(extract_people_from_text(text, label, url, allow_unseparated=(len(role_hits) == 1 and len(text) <= 240)))
        # BBB-style label/value where role precedes a name without punctuation.
        for match in re.finditer(
            rf"\b((?i:Owner|Principal|President|CEO|Founder|Co-Founder|Managing Member|Manager|Partner|Director|Officer|Registered Agent|Business Management))\b\s+(?:(?i:Mr|Ms|Mrs|Dr)\.?\s+)?({PERSON_NAME_PATTERN})",
            text,
        ):
            role, person_name = normalize_role(match.group(1)), normalize_person_name(match.group(2))
            if is_plausible_person_name(person_name):
                people.append({
                    "name": person_name,
                    "role": role,
                    "corporate_relationship": role,
                    "direct_phones": [], "direct_emails": [], "associated_addresses": [],
                    "relatives_and_associates": [],
                    "evidence_sources": [source_ref(label, url)],
                })
    return people


def extract_people_search_primary(text: str, label: str, url: str) -> List[Dict[str, Any]]:
    people: List[Dict[str, Any]] = []
    patterns = [
        rf"(?:Name|Subscriber|Subscriber Name|Caller Name|Phone Owner|Owner Name)\s*[:\-]\s*({PERSON_NAME_PATTERN})",
        rf"(?:Associated Person|Primary Person)\s*[:\-]\s*({PERSON_NAME_PATTERN})",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text or "", re.I):
            name = normalize_person_name(match.group(1))
            if is_plausible_person_name(name):
                people.append({
                    "name": name, "role": "Public Record Subject", "corporate_relationship": "Public Record Subject",
                    "direct_phones": [], "direct_emails": [], "associated_addresses": [],
                    "relatives_and_associates": [], "evidence_sources": [source_ref(label, url)],
                })
    return people


def parse_numpi_carrier(text: str) -> Dict[str, str]:
    result: Dict[str, str] = {}
    carrier_match = re.search(r"\bCarrier\s*[:\-]\s*([^|\n]{2,80})", text or "", re.I)
    type_match = re.search(r"\b(?:Line Type|Phone Type|Type)\s*[:\-]\s*(Mobile|Wireless|Landline|Fixed|VoIP|Toll[- ]?Free)", text or "", re.I)
    if carrier_match:
        result["carrier"] = re.sub(r"\s+", " ", carrier_match.group(1)).strip()
    if type_match:
        value = type_match.group(1).lower()
        result["line_type"] = "Mobile" if value in {"mobile", "wireless"} else "Landline" if value in {"landline", "fixed"} else "VoIP" if value == "voip" else "Toll-Free"
    return result


def extract_relatives(text: str, label: str, url: str) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    rel_pattern = "|".join(re.escape(x) for x in sorted(RELATIONSHIP_LABELS, key=len, reverse=True))
    rel_group = rf"(?i:{rel_pattern})"
    person = rf"({PERSON_TOKEN_PATTERN}(?:\s+{PERSON_TOKEN_PATTERN}){{1,3}}?)"
    boundary = rf"(?=\s+(?:{rel_group})\b|\s+(?i:phone|address|email|age|current|previous|lives)\b|\s+\d{{1,5}}\b|$)"
    pattern = re.compile(rf"\b({rel_group})\b\s*[:\-]?\s*{person}{boundary}")
    for match in pattern.finditer(text or ""):
        relationship = match.group(1).title()
        if relationship.lower() in {"relatives", "family", "family member"}:
            relationship = "Relative"
        elif relationship.lower() in {"associates", "business associates"}:
            relationship = "Associate" if relationship.lower() == "associates" else "Business Associate"
        name = normalize_person_name(match.group(2))
        if is_plausible_person_name(name):
            results.append({"name": name, "relationship": relationship, "sources": [source_ref(label, url)]})
    return results


async def extract_scoped_person_contacts(html: str, label: str, url: str, target_domain: str, verifier: ContactVerifier) -> List[Dict[str, Any]]:
    parser = LexborHTMLParser(html or "")
    role_rx = re.compile(r"\b(owner|principal|president|ceo|chief executive officer|founder|co-founder|managing member|manager|partner|director|officer|registered agent)\b", re.I)
    merged: List[Dict[str, Any]] = []
    seen_blocks: Set[str] = set()
    for node in parser.root.css("article, section, li, tr, .card, .profile, .team-member, [class*='contact'], [class*='management']"):
        block = re.sub(r"\s+", " ", node.text(separator=" ")).strip()
        if not block or len(block) > 1400 or not role_rx.search(block):
            continue
        fingerprint = block[:300].lower()
        if fingerprint in seen_blocks:
            continue
        seen_blocks.add(fingerprint)
        people = extract_people_from_text(block, label, url, allow_unseparated=(len(role_rx.findall(block)) == 1))
        if not people:
            people = extract_labeled_people(node.html or "", label, url)
        unique_people: List[Dict[str, Any]] = []
        for person in people:
            if not any(person_key(p.get("name", "")) == person_key(person.get("name", "")) for p in unique_people):
                unique_people.append(person)
        if not unique_people:
            continue
        block_html = node.html or ""
        block_phones = extract_phones(block_html, block, label, url)
        block_emails = await extract_emails(block_html, block, target_domain, label, url, verifier)
        block_addresses = extract_addresses(block_html, block, label, url)
        for person in unique_people:
            parts = normalize_person_name(person.get("name", "")).lower().split()
            if len(unique_people) == 1:
                merge_record_list(person.setdefault("direct_phones", []), block_phones, "number")
                merge_record_list(person.setdefault("associated_addresses", []), block_addresses, "address")
            if len(parts) >= 2:
                first, last = parts[0], parts[-1]
                patterns = {first, first + last, first[0] + last, first + last[0]}
                for email in block_emails:
                    local = email.get("email", "").split("@", 1)[0].lower().replace(".", "").replace("_", "").replace("-", "")
                    if local in patterns or len(unique_people) == 1:
                        merge_record_list(person.setdefault("direct_emails", []), [email], "email")
            merged.append(person)
    return merged


async def extract_page_profile(html: str, url: str, label: str, target_domain: str, verifier: ContactVerifier) -> Dict[str, Any]:
    parser = LexborHTMLParser(html or "")
    visible = trafilatura.extract(html or "") or parser.root.text(separator=" ")
    visible = re.sub(r"\s+", " ", visible).strip()
    phones = extract_phones(html, visible, label, url)
    emails = await extract_emails(html, visible, target_domain, label, url, verifier)
    addresses = extract_addresses(html, visible, label, url)
    schema_people, schema_addresses = parse_jsonld_people_addresses(html, label, url)
    people = extract_people_from_text(visible, label, url) + extract_labeled_people(html, label, url) + schema_people
    people.extend(await extract_scoped_person_contacts(html, label, url, target_domain, verifier))
    if label in {"PeopleSearchNow", "USPhoneBook", "TruePeopleSearch", "FastPeopleSearch", "That'sThem", "NumLookup", "Sync.ME", "SpyDialer", "RocketReach"}:
        people.extend(extract_people_search_primary(visible, label, url))
        for heading in parser.root.css("h1, h2"):
            heading_text = re.sub(r"\s+", " ", heading.text(separator=" ")).strip()
            if re.fullmatch(PERSON_NAME_PATTERN, heading_text) and not any(w in heading_text.lower() for w in ("phone search", "people search", "search results", "reverse lookup")):
                people.append({
                    "name": normalize_person_name(heading_text), "role": "Public Record Subject", "corporate_relationship": "Public Record Subject",
                    "direct_phones": [], "direct_emails": [], "associated_addresses": [], "relatives_and_associates": [],
                    "evidence_sources": [source_ref(label, url)],
                })
                break
    addresses.extend(schema_addresses)
    socials = extract_socials(html, label, url)
    relatives = extract_relatives(visible, label, url)
    return {
        "phones": phones,
        "emails": emails,
        "addresses": addresses,
        "people": people,
        "socials": socials,
        "relatives": relatives,
        "text": visible,
    }


# ---------------------------------------------------------------------------
# Search discovery and source adapters
# ---------------------------------------------------------------------------

@dataclass
class SourceTarget:
    label: str
    url: str
    kind: str = "directory"


class MasterMultiSourceEngine:
    def __init__(self) -> None:
        self.db = DB

    async def _emit_fetch_status(self, result: FetchResult, label: str) -> str:
        if result.event == "SOURCE_COMPLETED":
            return await sse_event("SOURCE_COMPLETED", f"{label} responded in {result.duration:.2f}s.", {"source": label, "url": result.final_url, "status": result.status, "duration": result.duration})
        if result.event == "RATE_LIMITED_429":
            return await sse_event("RATE_LIMITED_429", f"{label} rate-limited the request (HTTP 429).", {"source": label, "url": result.final_url, "status": result.status})
        if result.event == "CHALLENGE_DETECTED":
            return await sse_event("CHALLENGE_DETECTED", f"{label} presented a security/challenge page.", {"source": label, "url": result.final_url, "status": result.status})
        if result.event == "SOURCE_BLOCKED_403":
            return await sse_event("SOURCE_BLOCKED_403", f"{label} blocked the request (HTTP 403).", {"source": label, "url": result.final_url, "status": result.status})
        return await sse_event("SOURCE_FAILED", f"{label} could not be read: {result.error or 'request failed'}", {"source": label, "url": result.final_url, "status": result.status})

    async def search_index(self, query: str, network: NetworkClient) -> Tuple[List[str], List[Dict[str, Any]]]:
        urls: List[str] = []
        audit: List[Dict[str, Any]] = []
        seen: Set[str] = set()
        search_hosts = ("bing.com", "duckduckgo.com", "google.com", "yahoo.com")

        def accept(raw: str) -> None:
            candidate = clean_search_url(html_lib.unescape(raw or "").replace("\\/", "/"))
            candidate = candidate.strip(" \t\r\n'\"<>()[]{}.,")
            if not candidate.startswith(("http://", "https://")):
                return
            host = hostname_from_url(candidate)
            if not host or any(search_host in host for search_host in search_hosts):
                return
            if candidate not in seen:
                seen.add(candidate)
                urls.append(candidate)

        for engine_name, template in SEARCH_ENGINES:
            search_url = template.format(q=quote_plus(query))
            result = await network.fetch(search_url, allow_browser=False, validate_ssrf=False)
            audit.append({"engine": engine_name, "event": result.event, "status": result.status, "duration": result.duration})
            if result.event != "SOURCE_COMPLETED" or not result.html:
                continue

            raw_html = result.html
            parser = LexborHTMLParser(raw_html)
            for a in parser.root.css("a[href]"):
                accept(a.attributes.get("href", ""))

            # Bing RSS exposes canonical result links in <link> elements.
            if engine_name == "Bing RSS":
                for match in re.findall(r"<link>\s*(https?://[^<]+)\s*</link>", raw_html, flags=re.I):
                    accept(match)

            # Fallback for search markup that hides result URLs in attributes or encoded payloads.
            for match in re.findall(r"https?://[^\s\"'<>]+", raw_html):
                accept(match)
            for match in re.findall(r"https?%3A%2F%2F[A-Za-z0-9%._~:/?#\[\]@!$&'()*+,;=\-]+", raw_html, flags=re.I):
                accept(unquote(match))

            if urls:
                break
        return urls[:24], audit

    async def discover_official_website(self, tokens: Dict[str, Any], network: NetworkClient) -> str:
        if tokens.get("domain"):
            return "https://" + tokens["domain"]
        company = tokens.get("company_name", "")
        state = tokens.get("state", "")
        urls, _ = await self.search_index(f'"{company}" {state} official website', network)
        blocked_hosts = {
            "bbb.org", "dnb.com", "yelp.com", "facebook.com", "linkedin.com", "instagram.com",
            "peoplesearchnow.com", "usphonebook.com", "truepeoplesearch.com", "fastpeoplesearch.com",
            "thatsthem.com", "numlookup.com", "sync.me", "spydialer.com", "numpi.com",
        }
        for url in urls:
            host = hostname_from_url(url)
            root = registered_domain(host)
            if root in blocked_hosts:
                continue
            first = host.split(".", 1)[0]
            if first in ASSET_SUBDOMAINS:
                continue
            return f"https://{root}"
        return ""

    async def discover_directory(self, label: str, domain: str, query: str, network: NetworkClient) -> str:
        variations = [
            f'site:{domain} {query}',
            f'{query} "{label}"',
        ]
        if label == "BBB":
            variations.insert(0, f'site:bbb.org/us/ {query}')
        elif label == "D&B":
            variations.insert(0, f'site:dnb.com/business-directory/company-profiles {query}')
        seen: Set[str] = set()
        for discovery_query in variations:
            urls, _ = await self.search_index(discovery_query, network)
            for url in urls:
                clean = clean_search_url(url)
                if clean in seen:
                    continue
                seen.add(clean)
                host = hostname_from_url(clean)
                if host == domain or host.endswith("." + domain):
                    return clean
        return ""

    async def crawl_subpages(self, root_url: str, root_html: str, target_domain: str, network: NetworkClient, verifier: ContactVerifier) -> AsyncGenerator[Tuple[str, Dict[str, Any]], None]:
        parser = LexborHTMLParser(root_html or "")
        links: List[str] = []
        seen: Set[str] = set()
        for a in parser.root.css("a[href]"):
            href = a.attributes.get("href", "")
            if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
                continue
            full = urljoin(root_url, href)
            host = hostname_from_url(full)
            if not domain_matches(host, target_domain):
                continue
            path = urlparse(full).path.lower()
            if any(k in path for k in ("about", "contact", "team", "leadership", "staff", "management", "executive")):
                if full not in seen:
                    seen.add(full)
                    links.append(full)
            if len(links) >= 6:
                break
        for sub_url in links:
            result = await network.fetch(sub_url)
            yield "fetch", {"label": f"Subpage {urlparse(sub_url).path}", "result": result}
            if result.event == "SOURCE_COMPLETED":
                profile = await extract_page_profile(result.html, result.final_url, "Official Website Subpage", target_domain, verifier)
                yield "profile", profile

    async def enrich_phone_people_sources(self, phone: Dict[str, Any], network: NetworkClient, verifier: ContactVerifier, target_domain: str) -> AsyncGenerator[Tuple[str, Dict[str, Any]], None]:
        digits = phone.get("digits") or phone_key(phone.get("number", ""))
        if len(digits) != 10:
            return
        hyphen = f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
        targets = [
            SourceTarget("PeopleSearchNow", f"https://www.peoplesearchnow.com/phone/{hyphen}", "people"),
            SourceTarget("USPhoneBook", f"https://www.usphonebook.com/phone-search/{hyphen}", "people"),
            SourceTarget("TruePeopleSearch", f"https://www.truepeoplesearch.com/resultphone?phoneno={digits}", "people"),
            SourceTarget("FastPeopleSearch", f"https://www.fastpeoplesearch.com/{hyphen}", "people"),
            SourceTarget("That'sThem", f"https://thatsthem.com/phone/{hyphen}", "people"),
            SourceTarget("NumLookup", f"https://www.numlookup.com/lookup?phone={digits}", "people"),
            SourceTarget("Sync.ME", f"https://sync.me/search/?number={digits}", "people"),
            SourceTarget("SpyDialer", f"https://spydialer.com/search?q={digits}", "people"),
            SourceTarget("Numpi", f"https://numpi.com/phone-info/{digits}/", "carrier"),
        ]
        for target in targets:
            result = await network.fetch(target.url)
            yield "fetch", {"label": target.label, "result": result}
            if result.event != "SOURCE_COMPLETED":
                continue
            profile = await extract_page_profile(result.html, result.final_url, target.label, target_domain, verifier)
            profile["seed_phone"] = digits
            if target.kind == "carrier":
                yield "carrier", {"digits": digits, "metadata": parse_numpi_carrier(profile.get("text", "")), "source": source_ref(target.label, result.final_url)}
            else:
                yield "profile", profile

    def merge_profile(self, dossier: Dict[str, Any], profile: Dict[str, Any], label: str, url: str, *, person_source: bool = False) -> None:
        entity = dossier["entity"]
        if person_source:
            people = profile.get("people", [])
            # People-search pages usually have a primary subscriber. Attach discovered records only when a person is explicitly present.
            if people:
                primary = people[0]
                primary.setdefault("direct_phones", [])
                primary.setdefault("direct_emails", [])
                primary.setdefault("associated_addresses", [])
                seed_digits = str(profile.get("seed_phone") or "")
                if len(seed_digits) == 10:
                    seed_record = next((json.loads(json.dumps(x)) for x in entity.get("phones", []) if x.get("digits") == seed_digits), None)
                    if seed_record is None:
                        seed_record = classify_phone(seed_digits, label, url)
                    if seed_record:
                        merge_evidence(seed_record.setdefault("sources", []), [source_ref(label, url)])
                        merge_record_list(primary["direct_phones"], [seed_record], "number")
                merge_record_list(primary["direct_phones"], profile.get("phones", []), "number")
                merge_record_list(primary["direct_emails"], profile.get("emails", []), "email")
                merge_record_list(primary["associated_addresses"], profile.get("addresses", []), "address")
                rels = primary.setdefault("relatives_and_associates", [])
                for rel in profile.get("relatives", []):
                    if not any(person_key(x.get("name", "")) == person_key(rel.get("name", "")) and x.get("relationship") == rel.get("relationship") for x in rels):
                        rels.append(rel)
                merge_person(dossier["people"], primary)
                for extra in people[1:]:
                    merge_person(dossier["people"], extra)
            return

        generic_emails, direct_emails = [], []
        for email in profile.get("emails", []):
            local = email.get("email", "").split("@", 1)[0].lower()
            (generic_emails if local in GENERIC_EMAIL_PREFIXES else direct_emails).append(email)
        merge_record_list(entity["phones"], profile.get("phones", []), "number")
        merge_record_list(entity["emails"], generic_emails, "email")
        merge_record_list(entity["addresses"], profile.get("addresses", []), "address")
        for social in profile.get("socials", []):
            if not any(x.get("url") == social.get("url") for x in entity["socials"]):
                entity["socials"].append(social)
        for person in profile.get("people", []):
            merged = merge_person(dossier["people"], person)
            # Associate a work inbox to a named person when local-part clearly matches their name.
            first_last = normalize_person_name(merged.get("name", "")).lower().split()
            if len(first_last) >= 2:
                first, last = first_last[0], first_last[-1]
                for email in direct_emails:
                    local = email.get("email", "").split("@", 1)[0].lower().replace(".", "").replace("_", "")
                    candidates = {first + last, first[0] + last, first + last[0], first}
                    if local in candidates:
                        merge_record_list(merged.setdefault("direct_emails", []), [email], "email")
        # Direct emails not safely attributable stay at company level without claiming ownership.
        merge_record_list(entity["emails"], [e for e in direct_emails if not any(email_key(e["email"]) == email_key(pe.get("email", "")) for p in dossier["people"] for pe in p.get("direct_emails", []))], "email")

        src = source_ref(label, url)
        merge_evidence(dossier["sources"], [src])

    async def capture_source_snapshot(self, dossier: Dict[str, Any], label: str, url: str, proxy: str) -> Tuple[str, str]:
        metrics = dossier.setdefault("audit", {}).setdefault("snapshot_requests", {"attempted": 0, "captured": 0, "failed": 0})
        if len(dossier.setdefault("snapshots", [])) >= MAX_SNAPSHOTS_PER_DOSSIER:
            return "SNAPSHOT_SKIPPED", "Snapshot limit reached."
        if any(x.get("source_url") == url for x in dossier["snapshots"]):
            return "SNAPSHOT_SKIPPED", "Snapshot already captured for this source."
        metrics["attempted"] += 1
        snapshot, error = await BROWSER_POOL.capture_snapshot(url, proxy, label, dossier.get("cache_key", ""))
        if snapshot:
            dossier["snapshots"].append(snapshot)
            metrics["captured"] += 1
            return "SNAPSHOT_CAPTURED", f"Captured page preview for {label}."
        metrics["failed"] += 1
        return "SNAPSHOT_UNAVAILABLE", f"Page preview unavailable for {label}: {error or 'unknown error'}"

    async def run_master_pipeline(self, raw_query: str, custom_urls_json: str = "[]", refresh: bool = False) -> AsyncGenerator[str, None]:
        started = time.perf_counter()
        try:
            tokens = UniversalQueryParser.parse(raw_query)
        except ValueError as exc:
            yield await sse_event("ERROR", str(exc))
            return

        try:
            raw_custom = json.loads(custom_urls_json or "[]")
            if not isinstance(raw_custom, list):
                raw_custom = []
        except json.JSONDecodeError:
            raw_custom = []
        custom_urls: List[str] = []
        for value in raw_custom[:MAX_CUSTOM_URLS]:
            if isinstance(value, str) and value.strip():
                try:
                    custom_urls.append(validate_external_http_url(value.strip()))
                except ValueError:
                    continue

        proxy = self.db.get_setting("proxy_url", "")
        ttl_days = clamp_int(self.db.get_setting("cache_ttl_days", str(DEFAULT_CACHE_TTL_DAYS)), DEFAULT_CACHE_TTL_DAYS, 1, 90)
        smtp_checks = self.db.get_setting("smtp_checks", "1") != "0"
        max_concurrency = clamp_int(self.db.get_setting("max_concurrency", str(DEFAULT_MAX_CONCURRENCY)), DEFAULT_MAX_CONCURRENCY, 1, 12)
        max_phone_seeds = clamp_int(self.db.get_setting("max_phone_seeds", "4"), 4, 1, 20)
        snapshots_enabled = self.db.get_setting("snapshots_enabled", "1") != "0"
        cache_key = self.db.cache_key(tokens["normalized_query"], custom_urls, proxy)

        if not refresh:
            cached = self.db.get_cached(cache_key)
            if cached:
                cached.setdefault("ui_state", {})["lastQuery"] = cached.get("query", raw_query)
                yield await sse_event("CACHE_HIT", "Loaded a fresh cached dossier.", cached)
                return

        dossier = make_empty_dossier(tokens, cache_key)
        verifier = ContactVerifier(smtp_checks=smtp_checks)
        yield await sse_event("START", f"Research started for {tokens['company_name'] or tokens['raw']}.", {"cache_key": cache_key})

        async with NetworkClient(proxy=proxy, max_concurrency=max_concurrency) as network:
            # Explicit direct inputs are preserved but never pre-labeled verified.
            for raw_phone in tokens.get("phones", []):
                phone = classify_phone(raw_phone, "User Query", "")
                if phone:
                    merge_record_list(dossier["entity"]["phones"], [phone], "number")
            for raw_email in tokens.get("emails", []):
                status = await verifier.verify_email(raw_email, tokens.get("domain", ""))
                status["sources"] = [source_ref("User Query", "")]
                merge_record_list(dossier["entity"]["emails"], [status], "email")
            if tokens.get("owner"):
                merge_person(dossier["people"], {
                    "name": tokens["owner"], "role": "Owner/Executive Query", "corporate_relationship": "Query Target",
                    "direct_phones": [], "direct_emails": [], "associated_addresses": [], "relatives_and_associates": [],
                    "evidence_sources": [source_ref("User Query", "")],
                })

            official = await self.discover_official_website(tokens, network)
            target_domain = tokens.get("domain") or registered_domain(hostname_from_url(official))
            dossier["entity"]["official_domain"] = target_domain
            dossier["entity"]["official_website"] = official
            if official:
                yield await sse_event("ROUTE", f"Connecting to official website: {hostname_from_url(official)}")
                result = await network.fetch(official)
                yield await self._emit_fetch_status(result, "Official Website")
                if result.event == "SOURCE_COMPLETED":
                    profile = await extract_page_profile(result.html, result.final_url, "Official Website", target_domain, verifier)
                    self.merge_profile(dossier, profile, "Official Website", result.final_url)
                    if snapshots_enabled:
                        snap_type, snap_message = await self.capture_source_snapshot(dossier, "Official Website", result.final_url, proxy)
                        yield await sse_event(snap_type, snap_message)
                    async for kind, payload in self.crawl_subpages(result.final_url, result.html, target_domain, network, verifier):
                        if kind == "fetch":
                            yield await self._emit_fetch_status(payload["result"], payload["label"])
                        else:
                            self.merge_profile(dossier, payload, "Official Website Subpage", official)

            # High-yield corporate directories retained for Phase 1.
            company_query = f'"{tokens["company_name"]}" {tokens["state"]}'.strip()
            directory_specs = (("BBB", "bbb.org"), ("D&B", "dnb.com"))
            for label, domain in directory_specs:
                src_url = await self.discover_directory(label, domain, company_query, network)
                if not src_url:
                    yield await sse_event("SOURCE_SKIPPED", f"No matching {label} profile was discovered.")
                    continue
                result = await network.fetch(src_url)
                yield await self._emit_fetch_status(result, label)
                if result.event == "SOURCE_COMPLETED":
                    profile = await extract_page_profile(result.html, result.final_url, label, target_domain, verifier)
                    self.merge_profile(dossier, profile, label, result.final_url)
                    if snapshots_enabled:
                        snap_type, snap_message = await self.capture_source_snapshot(dossier, label, result.final_url, proxy)
                        yield await sse_event(snap_type, snap_message)
                    text = profile.get("text", "")
                    if label == "BBB":
                        complaints = re.findall(r"\b(\d+)\s+complaints?\b", text, flags=re.I)
                        if complaints:
                            dossier["entity"]["metadata"]["bbb_complaints"] = max(int(x) for x in complaints)
                    if label == "D&B":
                        rev = re.search(r"Sales\s*(?:Revenue|Volume)?(?:\s*\(\$?[MBK]\))?\s*[:$]?\s*\$?([\d.,]+\s*(?:Million|Billion|M|B|K)?)", text, re.I)
                        emp = re.search(r"Employees?\s*(?:\(all sites\)|total)?\s*[:]?\s*([\d,]+)", text, re.I)
                        if rev:
                            dossier["entity"]["metadata"]["dnb_revenue"] = "$" + rev.group(1).strip()
                        if emp:
                            dossier["entity"]["metadata"]["dnb_employees"] = emp.group(1).strip()
                        dossier.setdefault("source_index", {})["dnb"] = source_ref("D&B", result.final_url)

            # Custom direct targets receive the same complete extraction pipeline.
            for custom_url in custom_urls:
                result = await network.fetch(custom_url)
                yield await self._emit_fetch_status(result, "Custom Target")
                if result.event == "SOURCE_COMPLETED":
                    profile = await extract_page_profile(result.html, result.final_url, "Custom Target", target_domain, verifier)
                    self.merge_profile(dossier, profile, "Custom Target", result.final_url)
                    if snapshots_enabled:
                        snap_type, snap_message = await self.capture_source_snapshot(dossier, "Custom Target", result.final_url, proxy)
                        yield await sse_event(snap_type, snap_message)

            # Reverse-phone / people source fan-out for discovered numbers. Cap unique seeds to avoid runaway load.
            seed_phones: List[Dict[str, Any]] = []
            seen_seed: Set[str] = set()
            for phone in dossier["entity"]["phones"]:
                digits = phone.get("digits", "")
                if digits and digits not in seen_seed:
                    seen_seed.add(digits)
                    seed_phones.append(phone)
                if len(seed_phones) >= max_phone_seeds:
                    break
            for phone in seed_phones:
                yield await sse_event("PEOPLE_ROUTE", f"Checking public phone records for {phone.get('number', '')}.")
                async for kind, payload in self.enrich_phone_people_sources(phone, network, verifier, target_domain):
                    if kind == "fetch":
                        yield await self._emit_fetch_status(payload["result"], payload["label"])
                    elif kind == "carrier":
                        for stored in dossier["entity"]["phones"]:
                            if stored.get("digits") == payload.get("digits"):
                                for field_name, field_value in payload.get("metadata", {}).items():
                                    if field_value:
                                        stored[field_name] = field_value
                                merge_evidence(stored.setdefault("sources", []), [payload["source"]])
                    else:
                        self.merge_profile(dossier, payload, "People Source", "", person_source=True)

            # RocketReach is used as a targeted executive discovery source, not as a generic directory card.
            for person in dossier["people"][:3]:
                person_name = person.get("name", "")
                if not person_name:
                    continue
                rr_urls, _ = await self.search_index(f'site:rocketreach.co "{person_name}" "{tokens["company_name"]}"', network)
                rr_url = next((u for u in rr_urls if hostname_from_url(u).endswith("rocketreach.co")), "")
                if not rr_url:
                    continue
                rr_result = await network.fetch(rr_url)
                yield await self._emit_fetch_status(rr_result, "RocketReach")
                if rr_result.event == "SOURCE_COMPLETED":
                    rr_profile = await extract_page_profile(rr_result.html, rr_result.final_url, "RocketReach", target_domain, verifier)
                    self.merge_profile(dossier, rr_profile, "RocketReach", rr_result.final_url, person_source=True)

            dossier["audit"]["requests"] = network.metrics.as_dict()

        dossier["audit"]["completed_at"] = iso_now()
        dossier["audit"]["elapsed_seconds"] = round(time.perf_counter() - started, 2)
        dossier.setdefault("ui_state", {})["lastQuery"] = raw_query
        self.db.save_dossier(dossier, ttl_days)
        yield await sse_event(
            "COMPLETE",
            f"Research complete: {dossier['audit']['requests']['attempted']} real network requests attempted in {dossier['audit']['elapsed_seconds']:.2f}s.",
            dossier,
        )

    async def scrape_and_merge(self, raw_url: str, entity_name: str, cache_key: str = "") -> AsyncGenerator[str, None]:
        try:
            safe_url = await SSRFGuard.validate(raw_url)
        except ValueError as exc:
            yield await sse_event("SSRF_BLOCKED", str(exc))
            return
        dossier: Optional[Dict[str, Any]] = None
        if cache_key:
            dossier = self.db.get_cached(cache_key)
        if dossier is None and entity_name:
            dossier = self.db.latest_for_entity(entity_name)
        if dossier is None:
            tokens = UniversalQueryParser.parse(entity_name or hostname_from_url(safe_url))
            proxy = self.db.get_setting("proxy_url", "")
            cache_key = self.db.cache_key(tokens["normalized_query"], [safe_url], proxy)
            dossier = make_empty_dossier(tokens, cache_key)
        target_domain = dossier.get("entity", {}).get("official_domain") or registered_domain(hostname_from_url(safe_url))
        proxy = self.db.get_setting("proxy_url", "")
        smtp_checks = self.db.get_setting("smtp_checks", "1") != "0"
        ttl_days = clamp_int(self.db.get_setting("cache_ttl_days", str(DEFAULT_CACHE_TTL_DAYS)), DEFAULT_CACHE_TTL_DAYS, 1, 90)
        verifier = ContactVerifier(smtp_checks=smtp_checks)
        async with NetworkClient(proxy=proxy) as network:
            yield await sse_event("START_DIRECT", f"Scrape & Merge started for {hostname_from_url(safe_url)}.")
            result = await network.fetch(safe_url)
            yield await self._emit_fetch_status(result, "Direct URL")
            if result.event != "SOURCE_COMPLETED":
                return
            profile = await extract_page_profile(result.html, result.final_url, "Custom Target", target_domain, verifier)
            self.merge_profile(dossier, profile, "Custom Target", result.final_url)
            if self.db.get_setting("snapshots_enabled", "1") != "0":
                snap_type, snap_message = await self.capture_source_snapshot(dossier, "Custom Target", result.final_url, proxy)
                yield await sse_event(snap_type, snap_message)
            dossier["audit"]["requests"] = network.metrics.as_dict()
            dossier.setdefault("ui_state", {})["lastQuery"] = dossier.get("query", entity_name)
            self.db.save_dossier(dossier, ttl_days)
            yield await sse_event("MERGE_COMPLETE", "Direct scrape merged and persisted to SQLite.", dossier)


# ---------------------------------------------------------------------------
# SSE helpers and API security
# ---------------------------------------------------------------------------

async def sse_event(event_type: str, message: str, data: Optional[Dict[str, Any]] = None) -> str:
    payload = {
        "timestamp": iso_now(),
        "type": event_type,
        "message": message,
        "data": data or {},
    }
    return "data: " + json.dumps(payload, ensure_ascii=False) + "\n\n"


def is_allowed_local_origin(origin: Optional[str]) -> bool:
    if not origin:
        return True
    try:
        parsed = urlparse(origin)
    except ValueError:
        return False
    host = (parsed.hostname or "").lower()
    if host in {"localhost", "127.0.0.1", "::1"}:
        return True
    try:
        ip = ipaddress.ip_address(host)
    except ValueError:
        return False
    return ip.is_private


def require_local_origin(origin: Optional[str]) -> None:
    if not is_allowed_local_origin(origin):
        raise HTTPException(status_code=403, detail="This operation is restricted to local application origins")


SERVER: Any = None


@asynccontextmanager
async def lifespan(_: FastAPI):
    DB.initialize()
    yield
    for task in list(BULK_TASKS):
        if not task.done():
            task.cancel()
    if BULK_TASKS:
        await asyncio.gather(*list(BULK_TASKS), return_exceptions=True)
    await BROWSER_POOL.close()


app = FastAPI(
    title="Universal Intelligence & Contact Extraction Engine",
    version=ENGINE_VERSION,
    lifespan=lifespan,
)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
app.mount("/snapshots", StaticFiles(directory=SNAPSHOT_DIR), name="snapshots")
app.mount("/pwa", StaticFiles(directory=PWA_DIR, html=True), name="pwa")
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"^https?://(?:localhost|127\.0\.0\.1|10(?:\.\d{1,3}){3}|192\.168(?:\.\d{1,3}){2}|172\.(?:1[6-9]|2\d|3[01])(?:\.\d{1,3}){2})(?::\d+)?$",
    allow_credentials=False,
    allow_methods=["GET", "POST", "DELETE"],
    allow_headers=["Content-Type", "X-Requested-With"],
)


@app.get("/", response_class=HTMLResponse)
async def browser_ui() -> FileResponse:
    index_path = os.path.join(WEB_DIR, "index.html")
    if not os.path.isfile(index_path):
        raise HTTPException(status_code=503, detail="Browser frontend is missing")
    return FileResponse(index_path, media_type="text/html")


@app.get("/api/v1/health")
async def health() -> Dict[str, Any]:
    return {
        "ok": True,
        "engine_version": ENGINE_VERSION,
        "schema_version": SCHEMA_VERSION,
        "playwright": PLAYWRIGHT_OK,
        "dns": DNS_OK,
        "bulk": {"xlsx": OPENPYXL_OK, "pdf": PYPDF_OK, "ocr": OCR_OK},
        "pwa": os.path.isfile(os.path.join(PWA_DIR, "manifest.webmanifest")),
    }


@app.get("/api/v1/research/stream")
async def stream_search(query: str = Query(...), custom_urls: str = Query("[]"), refresh: bool = Query(False)) -> StreamingResponse:
    engine = MasterMultiSourceEngine()
    return StreamingResponse(
        engine.run_master_pipeline(query, custom_urls, refresh),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@app.get("/api/v1/research/url-stream")
async def stream_direct_url(
    url: str = Query(...), entity_name: str = Query(""), cache_key: str = Query("")
) -> StreamingResponse:
    engine = MasterMultiSourceEngine()
    return StreamingResponse(
        engine.scrape_and_merge(url, entity_name, cache_key),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@app.get("/api/v1/history")
async def get_history() -> List[Dict[str, Any]]:
    return DB.history()


@app.get("/api/v1/history/{history_id}")
async def get_history_item(history_id: int) -> Dict[str, Any]:
    dossier = DB.history_item(history_id)
    if dossier is None:
        raise HTTPException(status_code=404, detail="Dossier not found")
    dossier.setdefault("ui_state", {})["lastQuery"] = dossier.get("query", "")
    return dossier


@app.delete("/api/v1/history/{history_id}")
async def delete_history_item(history_id: int, origin: Optional[str] = Header(None)) -> Dict[str, Any]:
    require_local_origin(origin)
    if not DB.delete_history(history_id):
        raise HTTPException(status_code=404, detail="Dossier not found")
    return {"ok": True}


@app.delete("/api/v1/history")
async def clear_all_history(origin: Optional[str] = Header(None)) -> Dict[str, Any]:
    require_local_origin(origin)
    DB.clear_history()
    return {"ok": True}


@app.get("/api/v1/settings")
async def get_settings_endpoint() -> Dict[str, Any]:
    # No API-key fields are exposed or stored.
    proxy = DB.get_setting("proxy_url", "")
    safe_proxy = ""
    if proxy:
        parsed = urlparse(proxy)
        if parsed.hostname:
            safe_proxy = f"{parsed.scheme}://{parsed.hostname}:{parsed.port}" if parsed.port else f"{parsed.scheme}://{parsed.hostname}"
    return {
        "proxy_configured": bool(proxy),
        "proxy_display": safe_proxy,
        "cache_ttl_days": clamp_int(DB.get_setting("cache_ttl_days", str(DEFAULT_CACHE_TTL_DAYS)), DEFAULT_CACHE_TTL_DAYS, 1, 90),
        "smtp_checks": DB.get_setting("smtp_checks", "1") != "0",
        "max_concurrency": clamp_int(DB.get_setting("max_concurrency", str(DEFAULT_MAX_CONCURRENCY)), DEFAULT_MAX_CONCURRENCY, 1, 12),
        "max_phone_seeds": clamp_int(DB.get_setting("max_phone_seeds", "4"), 4, 1, 20),
        "snapshots_enabled": DB.get_setting("snapshots_enabled", "1") != "0",
    }


@app.post("/api/v1/settings")
async def save_settings_endpoint(payload: Dict[str, Any] = Body(...), origin: Optional[str] = Header(None)) -> Dict[str, Any]:
    require_local_origin(origin)
    allowed: Dict[str, str] = {}
    if "proxy_url" in payload:
        proxy = str(payload.get("proxy_url") or "").strip()
        if proxy:
            parsed = urlparse(proxy)
            if parsed.scheme not in {"http", "https", "socks5"} or not parsed.hostname:
                raise HTTPException(status_code=400, detail="Proxy must be http://, https://, or socks5:// with a valid host")
        allowed["proxy_url"] = proxy
    if "cache_ttl_days" in payload:
        allowed["cache_ttl_days"] = str(clamp_int(payload["cache_ttl_days"], DEFAULT_CACHE_TTL_DAYS, 1, 90))
    if "smtp_checks" in payload:
        allowed["smtp_checks"] = "1" if bool(payload["smtp_checks"]) else "0"
    if "max_concurrency" in payload:
        allowed["max_concurrency"] = str(clamp_int(payload["max_concurrency"], DEFAULT_MAX_CONCURRENCY, 1, 12))
    if "max_phone_seeds" in payload:
        allowed["max_phone_seeds"] = str(clamp_int(payload["max_phone_seeds"], 4, 1, 20))
    if "snapshots_enabled" in payload:
        allowed["snapshots_enabled"] = "1" if bool(payload["snapshots_enabled"]) else "0"
    DB.set_settings(allowed)
    return {"ok": True, "saved": sorted(allowed)}


@app.post("/api/v1/shutdown")
async def shutdown_server(origin: Optional[str] = Header(None)) -> Dict[str, Any]:
    require_local_origin(origin)
    if SERVER is None:
        return {"ok": False, "detail": "Server lifecycle is not controlled by the built-in runner"}
    SERVER.should_exit = True
    return {"ok": True, "detail": "Graceful shutdown requested"}


# ---------------------------------------------------------------------------
# Phase 2 bulk ingestion, exports, and snapshot controls
# ---------------------------------------------------------------------------

BULK_JOBS: Dict[str, Dict[str, Any]] = {}
BULK_TASKS: Set[asyncio.Task[Any]] = set()
BULK_QUEUE_SEMAPHORE = asyncio.Semaphore(1)

def _decode_tabular_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")

def _clean_bulk_row(row: Dict[str, Any]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for key, value in row.items():
        k = re.sub(r"\s+", " ", str(key or "")).strip()
        if not k:
            continue
        if value is None:
            out[k] = ""
        else:
            out[k] = re.sub(r"\s+", " ", str(value)).strip()
    return out

def _query_from_bulk_row(row: Dict[str, str]) -> str:
    lookup = {re.sub(r"[^a-z0-9]", "", k.lower()): v for k, v in row.items() if v}
    company_keys = ("company", "companyname", "business", "businessname", "legalname", "merchantname", "dba", "doingbusinessas")
    owner_keys = ("owner", "ownername", "contact", "contactname", "principal", "executive")
    state_keys = ("state", "businessstate", "companystate")
    direct_keys = ("website", "url", "domain", "email", "businessemail", "phone", "businessphone")
    company = next((lookup[k] for k in company_keys if lookup.get(k)), "")
    owner = next((lookup[k] for k in owner_keys if lookup.get(k)), "")
    state = next((lookup[k] for k in state_keys if lookup.get(k)), "")
    if company:
        parts = [company]
        if owner:
            parts.append(f"owner {owner}")
        if state:
            parts.append(f"in {state}")
        return " ".join(parts)[:MAX_QUERY_LENGTH]
    direct = next((lookup[k] for k in direct_keys if lookup.get(k)), "")
    if direct:
        return direct[:MAX_QUERY_LENGTH]
    values = [v for v in row.values() if v]
    return " ".join(values[:4])[:MAX_QUERY_LENGTH]

def _document_rows_from_text(text: str, source_label: str) -> List[Dict[str, str]]:
    clean = re.sub(r"[ \t]+", " ", text or "")
    labels = {
        "Business Name": r"(?:business|company|legal|merchant)\s*(?:name)?\s*[:#-]\s*([^\n\r]{2,100})",
        "DBA": r"(?:dba|doing business as)\s*[:#-]\s*([^\n\r]{2,100})",
        "Owner": r"(?:owner|principal|contact)\s*(?:name)?\s*[:#-]\s*([^\n\r]{2,80})",
        "Email": r"(?:email|e-mail)\s*[:#-]\s*([A-Za-z0-9_.+\-]+@(?:[A-Za-z0-9\-]+\.)+[A-Za-z]{2,63})",
        "Phone": r"(?:phone|telephone|mobile|cell)\s*[:#-]\s*((?:\+?1[-.\s]?)?\(?[2-9]\d{2}\)?[-.\s]?\d{3}[-.\s]?\d{4})",
        "State": r"(?:state)\s*[:#-]\s*([A-Z]{2}|[A-Za-z ]{4,20})",
    }
    row: Dict[str, str] = {"Source": source_label}
    for key, pattern in labels.items():
        match = re.search(pattern, clean, flags=re.I)
        if match:
            row[key] = re.sub(r"\s+", " ", match.group(1)).strip(" -:;,")
    if len(row) > 1:
        return [row]
    lines = [re.sub(r"\s+", " ", x).strip() for x in clean.splitlines() if len(x.strip()) > 2]
    candidate = next((x for x in lines if re.search(r"\b(?:LLC|INC|CORP|CORPORATION|COMPANY|CO\.|LTD|LP|LLP)\b", x, re.I)), "")
    if not candidate and lines:
        candidate = lines[0]
    return [{"Source": source_label, "Company": candidate[:160]}] if candidate else []

def parse_bulk_input(filename: str, data: bytes) -> List[Dict[str, str]]:
    if len(data) > MAX_BULK_FILE_BYTES:
        raise ValueError("File exceeds the 25 MB bulk-ingestion limit")
    suffix = Path(filename or "").suffix.lower()
    rows: List[Dict[str, str]] = []
    if suffix in {".csv", ".tsv", ".txt"}:
        text = _decode_tabular_bytes(data)
        delimiter = "\t" if suffix == ".tsv" or ("\t" in text.splitlines()[0] if text.splitlines() else False) else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        rows = [_clean_bulk_row(dict(r)) for r in reader if any(str(v or "").strip() for v in r.values())]
    elif suffix == ".xlsx":
        if not OPENPYXL_OK:
            raise ValueError("openpyxl is required for XLSX bulk files")
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers_raw = next(iterator, None)
        if not headers_raw:
            return []
        headers = [str(x or f"Column {i+1}").strip() for i, x in enumerate(headers_raw)]
        for values in iterator:
            raw = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
            if any(v not in (None, "") for v in raw.values()):
                rows.append(_clean_bulk_row(raw))
    elif suffix == ".pdf":
        if not PYPDF_OK:
            raise ValueError("pypdf is required for PDF bulk files")
        reader = PdfReader(io.BytesIO(data))
        for index, page in enumerate(reader.pages[:MAX_BULK_RECORDS], start=1):
            rows.extend(_document_rows_from_text(page.extract_text() or "", f"PDF page {index}"))
    elif suffix in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".bmp"}:
        if not OCR_OK:
            raise ValueError("Pillow and pytesseract are required for image OCR")
        try:
            text = pytesseract.image_to_string(Image.open(io.BytesIO(data)))
        except Exception as exc:
            raise ValueError(f"Local Tesseract OCR is unavailable: {str(exc)[:160]}") from exc
        rows = _document_rows_from_text(text, filename)
    else:
        raise ValueError("Supported bulk files: .xlsx, .csv, .tsv, .pdf, and common image formats")
    cleaned = [row for row in rows if _query_from_bulk_row(row)]
    return cleaned[:MAX_BULK_RECORDS]

async def _collect_dossier(query: str) -> Tuple[Optional[Dict[str, Any]], List[Dict[str, Any]]]:
    engine = MasterMultiSourceEngine()
    final: Optional[Dict[str, Any]] = None
    events: List[Dict[str, Any]] = []
    async for chunk in engine.run_master_pipeline(query, "[]", False):
        if not chunk.startswith("data: "):
            continue
        try:
            event = json.loads(chunk[6:].strip())
        except json.JSONDecodeError:
            continue
        events.append({"type": event.get("type", ""), "message": event.get("message", "")})
        if event.get("type") in {"COMPLETE", "CACHE_HIT"} and isinstance(event.get("data"), dict):
            final = event["data"]
    return final, events[-12:]

async def _run_bulk_job(job_id: str, rows: List[Dict[str, str]]) -> None:
    job = BULK_JOBS[job_id]
    try:
        async with BULK_QUEUE_SEMAPHORE:
            job["status"] = "running"
            job["started_at"] = iso_now()
            for index, row in enumerate(rows, start=1):
                query = _query_from_bulk_row(row)
                job["current"] = index
                job["current_query"] = query
                job["message"] = f"Processing lead {index} of {len(rows)}"
                try:
                    dossier, events = await _collect_dossier(query)
                    job["results"].append({"index": index, "input": row, "query": query, "dossier": dossier, "events": events, "error": "" if dossier else "No completed dossier"})
                    if dossier:
                        job["completed"] += 1
                    else:
                        job["failed"] += 1
                except Exception as exc:
                    job["results"].append({"index": index, "input": row, "query": query, "dossier": None, "events": [], "error": str(exc)[:300]})
                    job["failed"] += 1
                job["progress"] = round(index / max(1, len(rows)) * 100, 1)
            job["status"] = "completed" if job["failed"] == 0 else "completed_with_errors"
            job["message"] = f"Finished {len(rows)} leads: {job['completed']} completed, {job['failed']} failed."
    except asyncio.CancelledError:
        job["status"] = "cancelled"
        job["message"] = "Bulk job cancelled."
        raise
    finally:
        job["finished_at"] = iso_now()

def _bulk_export_rows(job: Dict[str, Any]) -> List[Dict[str, str]]:
    flat: List[Dict[str, str]] = []
    for result in job.get("results", []):
        dossier = result.get("dossier") or {}
        entity = dossier.get("entity", {})
        people = dossier.get("people", []) or [None]
        company_phones = "; ".join(x.get("number", "") for x in entity.get("phones", []))
        company_emails = "; ".join(x.get("email", "") for x in entity.get("emails", []))
        company_addresses = "; ".join(x.get("address", "") for x in entity.get("addresses", []))
        for person in people:
            p = person or {}
            sources = p.get("evidence_sources", []) if person else dossier.get("sources", [])
            flat.append({
                "input_index": str(result.get("index", "")),
                "input_query": result.get("query", ""),
                "company": entity.get("name", ""),
                "state": entity.get("state", ""),
                "official_domain": entity.get("official_domain", ""),
                "company_phones": company_phones,
                "company_emails": company_emails,
                "company_addresses": company_addresses,
                "person_name": p.get("name", ""),
                "role": p.get("role", ""),
                "roles": "; ".join(p.get("roles", [])),
                "direct_phones": "; ".join(x.get("number", "") for x in p.get("direct_phones", [])),
                "phone_types": "; ".join(x.get("line_type", "") for x in p.get("direct_phones", [])),
                "carriers": "; ".join(x.get("carrier", "") for x in p.get("direct_phones", [])),
                "direct_emails": "; ".join(x.get("email", "") for x in p.get("direct_emails", [])),
                "email_statuses": "; ".join(f"{x.get('mx_status','')}/{x.get('smtp_status','')}" for x in p.get("direct_emails", [])),
                "associated_addresses": "; ".join(x.get("address", "") for x in p.get("associated_addresses", [])),
                "relatives_associates": "; ".join(f"{x.get('name','')} ({x.get('relationship','')})" for x in p.get("relatives_and_associates", [])),
                "evidence_sources": "; ".join(sorted({x.get("label", "") for x in sources if x.get("label")})),
                "error": result.get("error", ""),
            })
    return flat

def _csv_safe(value: Any) -> str:
    text = str(value or "")
    stripped = text.lstrip(" \t\r\n")
    return "'" + text if stripped[:1] in {"=", "+", "-", "@"} else text

@app.post("/api/v1/bulk/jobs")
async def create_bulk_job(request: Request, filename: str = Query(...), origin: Optional[str] = Header(None)) -> Dict[str, Any]:
    require_local_origin(origin)
    data = await request.body()
    try:
        rows = await asyncio.to_thread(parse_bulk_input, filename, data)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not rows:
        raise HTTPException(status_code=400, detail="No searchable lead records were detected in the uploaded file")
    job_id = uuid.uuid4().hex[:16]
    BULK_JOBS[job_id] = {
        "id": job_id, "filename": Path(filename).name, "status": "queued", "total": len(rows), "current": 0,
        "completed": 0, "failed": 0, "progress": 0.0, "message": "Queued", "current_query": "",
        "created_at": iso_now(), "started_at": "", "finished_at": "", "results": [],
    }
    task = asyncio.create_task(_run_bulk_job(job_id, rows), name=f"bulk-{job_id}")
    BULK_TASKS.add(task)
    task.add_done_callback(BULK_TASKS.discard)
    # Keep memory bounded to recent jobs.
    if len(BULK_JOBS) > 20:
        for old_id in list(BULK_JOBS)[:-20]:
            if BULK_JOBS[old_id].get("status") not in {"queued", "running"}:
                BULK_JOBS.pop(old_id, None)
    return {"ok": True, "job_id": job_id, "total": len(rows)}

@app.get("/api/v1/bulk/jobs/{job_id}")
async def get_bulk_job(job_id: str) -> Dict[str, Any]:
    job = BULK_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Bulk job not found")
    summaries = []
    for item in job.get("results", [])[-25:]:
        dossier = item.get("dossier") or {}
        summaries.append({
            "index": item.get("index"),
            "query": item.get("query", ""),
            "company": dossier.get("entity", {}).get("name", ""),
            "people": len(dossier.get("people", [])),
            "phones": len(dossier.get("entity", {}).get("phones", [])) + sum(len(p.get("direct_phones", [])) for p in dossier.get("people", [])),
            "emails": len(dossier.get("entity", {}).get("emails", [])) + sum(len(p.get("direct_emails", [])) for p in dossier.get("people", [])),
            "error": item.get("error", ""),
        })
    return {k: v for k, v in job.items() if k != "results"} | {"recent_results": summaries}

@app.delete("/api/v1/bulk/jobs/{job_id}")
async def cancel_bulk_job(job_id: str, origin: Optional[str] = Header(None)) -> Dict[str, Any]:
    require_local_origin(origin)
    job = BULK_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Bulk job not found")
    for task in list(BULK_TASKS):
        if task.get_name() == f"bulk-{job_id}" and not task.done():
            task.cancel()
    return {"ok": True}

@app.get("/api/v1/bulk/jobs/{job_id}/export")
async def export_bulk_job(job_id: str, format: str = Query("csv", pattern="^(csv|xlsx)$")) -> Response:
    job = BULK_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Bulk job not found")
    rows = _bulk_export_rows(job)
    if not rows:
        raise HTTPException(status_code=409, detail="Bulk job has no exportable results yet")
    headers = list(rows[0].keys())
    if format == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: _csv_safe(v) for k, v in row.items()})
        payload = output.getvalue().encode("utf-8-sig")
        return Response(payload, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="bulk-{job_id}.csv"'})
    if not OPENPYXL_OK:
        raise HTTPException(status_code=503, detail="openpyxl is required for XLSX export")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Enriched Leads"
    sheet.append(headers)
    for row in rows:
        sheet.append([_csv_safe(row.get(k, "")) for k in headers])
    stream = io.BytesIO()
    workbook.save(stream)
    return Response(stream.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="bulk-{job_id}.xlsx"'})

@app.get("/api/v1/export/database")
async def export_database(origin: Optional[str] = Header(None)) -> Response:
    require_local_origin(origin)
    DB.initialize()
    fd, temp_path = tempfile.mkstemp(prefix="extractor-backup-", suffix=".sqlite")
    os.close(fd)
    try:
        source = sqlite3.connect(DB_PATH)
        target = sqlite3.connect(temp_path)
        try:
            source.backup(target)
        finally:
            target.close()
            source.close()
        payload = Path(temp_path).read_bytes()
    finally:
        with suppress(OSError):
            os.remove(temp_path)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return Response(payload, media_type="application/vnd.sqlite3", headers={"Content-Disposition": f'attachment; filename="extractor-backup-{stamp}.sqlite"'})

@app.post("/api/v1/snapshots/capture")
async def capture_snapshot_endpoint(payload: Dict[str, Any] = Body(...), origin: Optional[str] = Header(None)) -> Dict[str, Any]:
    require_local_origin(origin)
    raw_url = str(payload.get("url") or "").strip()
    label = str(payload.get("label") or "Source").strip()[:80]
    cache_key = str(payload.get("cache_key") or "").strip()
    if not raw_url:
        raise HTTPException(status_code=400, detail="Snapshot URL is required")
    try:
        safe_url = await SSRFGuard.validate(raw_url)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    dossier = DB.get_by_cache_key(cache_key) if cache_key else None
    proxy = DB.get_setting("proxy_url", "")
    target_key = cache_key or hashlib.sha256(safe_url.encode("utf-8")).hexdigest()
    snapshot, error = await BROWSER_POOL.capture_snapshot(safe_url, proxy, label, target_key)
    if snapshot is None:
        raise HTTPException(status_code=503, detail=error or "Snapshot capture unavailable")
    if dossier is not None:
        dossier.setdefault("snapshots", []).append(snapshot)
        # De-duplicate and keep the most recent configured maximum.
        by_url: Dict[str, Dict[str, Any]] = {}
        for item in dossier["snapshots"]:
            by_url[item.get("source_url", item.get("image_url", ""))] = item
        dossier["snapshots"] = list(by_url.values())[-MAX_SNAPSHOTS_PER_DOSSIER:]
        ttl_days = clamp_int(DB.get_setting("cache_ttl_days", str(DEFAULT_CACHE_TTL_DAYS)), DEFAULT_CACHE_TTL_DAYS, 1, 90)
        DB.save_dossier(dossier, ttl_days)
    return {"ok": True, "snapshot": snapshot}

# ---------------------------------------------------------------------------
# Explicit startup only; importing this module has no shortcut/port/process side effects.
# ---------------------------------------------------------------------------

APP_URL = "http://127.0.0.1:8000/"
HEALTH_URL = "http://127.0.0.1:8000/api/v1/health"


def _running_instance_is_ours(timeout: float = 1.2) -> bool:
    try:
        request = UrlRequest(HEALTH_URL, headers={"User-Agent": "IntelligenceExtractorLauncher/1.0"})
        with urlopen(request, timeout=timeout) as response:
            if response.status != 200:
                return False
            payload = json.loads(response.read().decode("utf-8", errors="replace"))
            return bool(payload.get("ok")) and str(payload.get("engine_version", "")).startswith("2026.08.21-phase2")
    except Exception:
        return False


def _port_8000_in_use() -> bool:
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(0.5)
    try:
        return sock.connect_ex(("127.0.0.1", 8000)) == 0
    finally:
        sock.close()


def _show_windows_error(message: str) -> None:
    if os.name != "nt":
        return
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(None, message, "Intelligence Extractor", 0x10)
    except Exception:
        pass


async def _open_browser_when_ready() -> None:
    for _ in range(120):
        if _running_instance_is_ours(timeout=0.6):
            webbrowser.open(APP_URL, new=2)
            return
        await asyncio.sleep(0.25)
    _show_windows_error("The Intelligence Extractor server did not become ready on 127.0.0.1:8000.")


async def _serve(open_browser: bool = True) -> None:
    global SERVER
    import uvicorn
    if open_browser:
        asyncio.create_task(_open_browser_when_ready())
    config = uvicorn.Config(app, host="0.0.0.0", port=8000, log_level="warning", access_log=False)
    SERVER = uvicorn.Server(config)
    await SERVER.serve()


def _main() -> int:
    args = set(sys.argv[1:])
    background = "--background" in args
    if "--shutdown" in args:
        try:
            request = UrlRequest(HEALTH_URL.replace("/health", "/shutdown"), data=b"{}", method="POST", headers={"Content-Type": "application/json"})
            with urlopen(request, timeout=2.5):
                pass
            return 0
        except Exception:
            return 0
    if _running_instance_is_ours():
        if not background:
            webbrowser.open(APP_URL, new=2)
        return 0
    if _port_8000_in_use():
        _show_windows_error("Port 8000 is already being used by another program. Intelligence Extractor did not start.")
        return 2
    try:
        asyncio.run(_serve(open_browser=not background))
        return 0
    except Exception as exc:
        _show_windows_error(f"Intelligence Extractor could not start.\n\n{exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(_main())
